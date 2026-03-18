"""
OnBrain AI - Data Indexing Service v2.1.0
LlamaIndex + VectorStore Pipeline uchun Service
Google Drive papkasidaq fayllarni indexlaydigan va Q&A qilishdagi qo'mak ko'rsatadigan modul
Fixed import issues for LlamaIndex 0.10.0+ compatibility (March 18, 2026)
"""

import logging
import os
import io
import json
import asyncio
from typing import Optional, Dict, List, Any, Tuple
from datetime import datetime

# LlamaIndex imports
from llama_index.core import VectorStoreIndex, Document
from llama_index.embeddings.openai import OpenAIEmbedding
from llama_index.core.vector_stores import SimpleVectorStore

# Google API
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import pandas as pd
import openpyxl

logger = logging.getLogger(__name__)


class DataIndexingService:
    """
    Google Drive papkasidagi ma'lumotlarni indexlaydigan va qidiruvchi xizmat
    LlamaIndex va VectorStore yordamida
    """
    
    def __init__(self, credentials: Credentials, openai_api_key: str):
        """
        Initialize indexing service
        
        Args:
            credentials: Google OAuth credentials
            openai_api_key: OpenAI API key for embeddings
        """
        self.credentials = credentials
        self.drive_service = build('drive', 'v3', credentials=credentials)
        self.sheets_service = build('sheets', 'v4', credentials=credentials)
        self.openai_api_key = openai_api_key
        
        # Vector store va index
        self.vector_store = SimpleVectorStore()
        self.index = None
        self.query_engine = None
        
        logger.info("✅ DataIndexingService initialized")
    
    async def index_folder_files(self, folder_id: str) -> Tuple[bool, str, int]:
        """
        Papka ichidagi barcha fayllarni indexlash
        
        Args:
            folder_id: Google Drive folder ID
            
        Returns:
            (success: bool, message: str, indexed_count: int)
        """
        try:
            logger.info(f"📂 Starting to index folder: {folder_id}")
            
            # 1. Papkada fayllarni topish
            documents = []
            indexed_count = 0
            
            # 1.1 Google Sheets fayllarini qidirish
            sheets_count = await self._index_google_sheets(folder_id, documents)
            indexed_count += sheets_count
            logger.info(f"📊 Indexed {sheets_count} Google Sheets")
            
            # 1.2 Excel fayllarini qidirish
            excel_count = await self._index_excel_files(folder_id, documents)
            indexed_count += excel_count
            logger.info(f"📄 Indexed {excel_count} Excel files")
            
            # 1.3 CSV fayllarini qidirish
            csv_count = await self._index_csv_files(folder_id, documents)
            indexed_count += csv_count
            logger.info(f"📋 Indexed {csv_count} CSV files")
            
            if not documents:
                return False, "📂 Papkada indexlanishi mumkin bo'lgan fayllar topilmadi", 0
            
            # 2. VectorStoreIndex yaratish
            logger.info(f"🔍 Creating VectorStoreIndex for {len(documents)} documents...")
            
            try:
                # Embedding model
                embed_model = OpenAIEmbedding(
                    api_key=self.openai_api_key,
                    model="text-embedding-3-small"
                )
                
                # Index yaratish
                self.index = VectorStoreIndex.from_documents(
                    documents,
                    embed_model=embed_model,
                    vector_store=self.vector_store,
                    show_progress=True
                )
                
                # Query engine yaratish
                self.query_engine = self.index.as_query_engine(
                    similarity_top_k=5,
                    response_mode="tree_summarize"
                )
                
                logger.info(f"✅ Index created successfully with {len(documents)} documents")
                return True, f"✅ {indexed_count} ta fayl muvaffaqiyatli indexlandi", indexed_count
                
            except Exception as e:
                logger.error(f"❌ Error creating index: {e}")
                return False, f"❌ Index yaratishda xato: {str(e)}", indexed_count
        
        except Exception as e:
            logger.error(f"❌ Error indexing folder: {e}", exc_info=True)
            return False, f"❌ Papka indexlashda xato: {str(e)}", 0
    
    async def _index_google_sheets(self, folder_id: str, documents: List[Document]) -> int:
        """Google Sheets fayllarini topib, indexlash"""
        try:
            logger.info(f"📊 Indexing Google Sheets from folder {folder_id}...")
            count = 0
            page_token = None
            
            query = f"'{folder_id}' in parents and mimeType='application/vnd.google-apps.spreadsheet' and trashed=false"
            
            while True:
                results = await asyncio.to_thread(
                    lambda: self.drive_service.files().list(
                        q=query,
                        spaces='drive',
                        pageSize=50,
                        pageToken=page_token,
                        fields='files(id, name)',
                        orderBy='name'
                    ).execute()
                )
                
                files = results.get('files', [])
                
                for file in files:
                    try:
                        sheet_id = file['id']
                        sheet_name = file['name']
                        
                        logger.info(f"   📄 Reading Google Sheet: {sheet_name}")
                        
                        # Spreadsheet ma'lumotlarini o'qish
                        sheet_data = await asyncio.to_thread(
                            lambda: self.sheets_service.spreadsheets().get(
                                spreadsheetId=sheet_id,
                                includeGridData=False
                            ).execute()
                        )
                        
                        # Barcha worksheet-larni o'qish
                        for sheet in sheet_data.get('sheets', []):
                            sheet_title = sheet['properties']['title']
                            
                            # Sheet ma'lumotlarini olish
                            values = await asyncio.to_thread(
                                lambda: self.sheets_service.spreadsheets().values().get(
                                    spreadsheetId=sheet_id,
                                    range=f"{sheet_title}!A:Z"
                                ).execute()
                            )
                            
                            rows = values.get('values', [])
                            if rows:
                                # Document yaratish
                                doc_text = f"📊 Sheet: {sheet_name} - {sheet_title}\n\n"
                                doc_text += self._format_table_for_index(rows)
                                
                                doc = Document(
                                    text=doc_text,
                                    metadata={
                                        'source': 'google_sheets',
                                        'sheet_name': sheet_name,
                                        'worksheet': sheet_title,
                                        'sheet_id': sheet_id,
                                        'rows': len(rows)
                                    }
                                )
                                documents.append(doc)
                                count += 1
                                logger.info(f"      ✅ Indexed {sheet_title} ({len(rows)} rows)")
                    
                    except Exception as e:
                        logger.warning(f"⚠️ Could not index sheet {file['name']}: {e}")
                        continue
                
                page_token = results.get('nextPageToken')
                if not page_token:
                    break
            
            return count
        
        except Exception as e:
            logger.error(f"❌ Error indexing Google Sheets: {e}", exc_info=True)
            return 0
    
    async def _index_excel_files(self, folder_id: str, documents: List[Document]) -> int:
        """Excel fayllarini topib, indexlash"""
        try:
            logger.info(f"📄 Indexing Excel files from folder {folder_id}...")
            count = 0
            page_token = None
            
            # Excel MIME types
            excel_mimes = [
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
                'application/vnd.ms-excel'  # .xls
            ]
            
            mime_query = " or ".join([f"mimeType='{mime}'" for mime in excel_mimes])
            query = f"'{folder_id}' in parents and ({mime_query}) and trashed=false"
            
            while True:
                results = await asyncio.to_thread(
                    lambda: self.drive_service.files().list(
                        q=query,
                        spaces='drive',
                        pageSize=50,
                        pageToken=page_token,
                        fields='files(id, name)',
                        orderBy='name'
                    ).execute()
                )
                
                files = results.get('files', [])
                
                for file in files:
                    try:
                        file_id = file['id']
                        file_name = file['name']
                        
                        logger.info(f"   📄 Downloading Excel file: {file_name}")
                        
                        # Faylni yuklab olish
                        request = self.drive_service.files().get_media(fileId=file_id)
                        file_content = io.BytesIO()
                        downloader = MediaIoBaseDownload(file_content, request)
                        
                        done = False
                        while not done:
                            status, done = await asyncio.to_thread(downloader.next_chunk)
                        
                        file_content.seek(0)
                        
                        # Excel faylni o'qish
                        excel_data = await self._read_excel_file(file_content, file_name)
                        
                        if excel_data:
                            doc = Document(
                                text=excel_data,
                                metadata={
                                    'source': 'excel',
                                    'file_name': file_name,
                                    'file_id': file_id
                                }
                            )
                            documents.append(doc)
                            count += 1
                            logger.info(f"      ✅ Indexed {file_name}")
                    
                    except Exception as e:
                        logger.warning(f"⚠️ Could not index Excel file {file['name']}: {e}")
                        continue
                
                page_token = results.get('nextPageToken')
                if not page_token:
                    break
            
            return count
        
        except Exception as e:
            logger.error(f"❌ Error indexing Excel files: {e}", exc_info=True)
            return 0
    
    async def _index_csv_files(self, folder_id: str, documents: List[Document]) -> int:
        """CSV fayllarini topib, indexlash"""
        try:
            logger.info(f"📋 Indexing CSV files from folder {folder_id}...")
            count = 0
            page_token = None
            
            query = f"'{folder_id}' in parents and mimeType='text/csv' and trashed=false"
            
            while True:
                results = await asyncio.to_thread(
                    lambda: self.drive_service.files().list(
                        q=query,
                        spaces='drive',
                        pageSize=50,
                        pageToken=page_token,
                        fields='files(id, name)',
                        orderBy='name'
                    ).execute()
                )
                
                files = results.get('files', [])
                
                for file in files:
                    try:
                        file_id = file['id']
                        file_name = file['name']
                        
                        logger.info(f"   📋 Downloading CSV file: {file_name}")
                        
                        # Faylni yuklab olish
                        request = self.drive_service.files().get_media(fileId=file_id)
                        file_content = io.BytesIO()
                        downloader = MediaIoBaseDownload(file_content, request)
                        
                        done = False
                        while not done:
                            status, done = await asyncio.to_thread(downloader.next_chunk)
                        
                        file_content.seek(0)
                        
                        # CSV faylni o'qish
                        df = pd.read_csv(file_content)
                        
                        doc_text = f"📋 CSV File: {file_name}\n\n"
                        doc_text += self._format_dataframe_for_index(df)
                        
                        doc = Document(
                            text=doc_text,
                            metadata={
                                'source': 'csv',
                                'file_name': file_name,
                                'file_id': file_id,
                                'rows': len(df)
                            }
                        )
                        documents.append(doc)
                        count += 1
                        logger.info(f"      ✅ Indexed {file_name} ({len(df)} rows)")
                    
                    except Exception as e:
                        logger.warning(f"⚠️ Could not index CSV file {file['name']}: {e}")
                        continue
                
                page_token = results.get('nextPageToken')
                if not page_token:
                    break
            
            return count
        
        except Exception as e:
            logger.error(f"❌ Error indexing CSV files: {e}", exc_info=True)
            return 0
    
    async def _read_excel_file(self, file_content: io.BytesIO, file_name: str) -> Optional[str]:
        """Excel faylni o'qib, text formatiga o'tkazish"""
        try:
            workbook = openpyxl.load_workbook(file_content, data_only=True)
            
            doc_text = f"📄 Excel File: {file_name}\n\n"
            
            for worksheet in workbook.sheetnames:
                ws = workbook[worksheet]
                doc_text += f"📑 Sheet: {worksheet}\n"
                
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(row)
                
                doc_text += self._format_table_for_index(rows)
                doc_text += "\n\n"
            
            return doc_text
        
        except Exception as e:
            logger.error(f"❌ Error reading Excel file: {e}")
            return None
    
    def _format_table_for_index(self, rows: List[List[Any]]) -> str:
        """Jadval-ni index uchun tekst formatiga o'tkazish"""
        if not rows:
            return "Jadval bo'sh"
        
        text = "| "
        
        # Header row
        if len(rows) > 0:
            header = [str(cell) if cell is not None else "" for cell in rows[0]]
            text += " | ".join(header) + " |\n"
            text += "|" + "|".join(["---"] * len(header)) + "|\n"
        
        # Data rows (birinchi 50 qatorni oling)
        for row in rows[1:min(51, len(rows))]:
            row_str = [str(cell) if cell is not None else "" for cell in row]
            text += "| " + " | ".join(row_str) + " |\n"
        
        if len(rows) > 50:
            text += f"\n... va {len(rows) - 50} ta ko'proq qator\n"
        
        return text
    
    def _format_dataframe_for_index(self, df: pd.DataFrame) -> str:
        """Pandas DataFrame-ni index uchun tekst formatiga o'tkazish"""
        if df.empty:
            return "Jadval bo'sh"
        
        # Birinchi 50 qatorni oling
        df_sample = df.head(50)
        
        text = df_sample.to_string()
        
        if len(df) > 50:
            text += f"\n\n... va {len(df) - 50} ta ko'proq qator"
        
        return text
    
    async def query_index(self, question: str) -> Tuple[bool, str]:
        """
        Index-dan savol-javob olish
        
        Args:
            question: Foydalanuvchining savoli (O'zbek tilida)
            
        Returns:
            (success: bool, answer: str)
        """
        try:
            if not self.query_engine:
                return False, "❌ Index hali tayyar emas. Iltimos, papka indexlashni kutib turing."
            
            logger.info(f"🔍 Querying index with: {question}")
            
            # Query bajarish
            response = await asyncio.to_thread(
                lambda: self.query_engine.query(question)
            )
            
            answer = str(response)
            logger.info(f"✅ Got answer: {answer[:100]}...")
            
            return True, answer
        
        except Exception as e:
            logger.error(f"❌ Error querying index: {e}", exc_info=True)
            return False, f"❌ Savol javobida xato: {str(e)}"
    
    def get_index_status(self) -> Dict[str, Any]:
        """Index-ning holati haqida ma'lumot"""
        if self.index is None:
            return {
                'status': 'not_created',
                'message': '❌ Index yaratilmagan'
            }
        
        vector_count = len(self.vector_store.data)
        
        return {
            'status': 'ready',
            'message': f'✅ Index tayyor ({vector_count} vectors)',
            'vector_count': vector_count
        }
