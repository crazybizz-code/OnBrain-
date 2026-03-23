# OnBrain AI Bot v2.2.0 - Koyeb Migration (March 22, 2026)
import asyncio
import io
import json
import logging
import os
import re
import sqlite3
from html import escape as html_escape
import secrets
import time
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any

import openpyxl
import xlrd
from aiohttp import web
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.exceptions import TelegramBadRequest, TelegramAPIError
from aiogram.filters import CommandStart, Command
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    WebAppInfo,
)
from dotenv import load_dotenv
from googleapiclient.discovery import build
from google.auth.transport.requests import Request as GoogleAuthRequest
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
import httpx

import gspread

# Import GitHub OAuth Service
from github_oauth import GitHubOAuthService

# Import Data Indexing Service (optional — graceful fallback if file missing)
try:
    from data_indexing_service import DataIndexingService
except ImportError:
    DataIndexingService = None  # type: ignore


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger("onbrain-ai-bot")

# ==================== VERSION & FEATURES ====================
BOT_VERSION = "2.1.0"
FEATURES = {
    "ai_qa": True,           # AI Q&A from indexed data
    "data_indexing": True,   # LlamaIndex VectorStore indexing
    "google_sheets": True,   # Google Sheets integration
    "google_drive": True,    # Google Drive folder reading
    "tavily_search": True,   # Web search via Tavily
    "grok_ai": True,         # Grok AI for spreadsheet Q&A
}
logger.info(f"🚀 OnBrain AI Bot v{BOT_VERSION} - Features: {FEATURES}")

# ==================== SECURITY UTILITIES ====================

class RateLimiter:
    """Rate limiting to prevent brute force attacks"""
    def __init__(self, max_requests: int = 10, time_window: int = 60):
        self.max_requests = max_requests
        self.time_window = time_window
        self.requests: dict[int, list[float]] = {}
    
    def is_allowed(self, user_id: int) -> bool:
        """Check if user is within rate limit"""
        now = time.time()
        if user_id not in self.requests:
            self.requests[user_id] = []
        
        # Remove old requests outside the time window
        self.requests[user_id] = [
            req_time for req_time in self.requests[user_id]
            if now - req_time < self.time_window
        ]
        
        if len(self.requests[user_id]) >= self.max_requests:
            logger.warning(f"⚠️ Rate limit exceeded for user {user_id}")
            return False
        
        self.requests[user_id].append(now)
        return True


class InputValidator:
    """Validate and sanitize user inputs"""
    
    @staticmethod
    def validate_email(email: str) -> bool:
        """Validate email format"""
        if not email or len(email) > 254:
            return False
        return EMAIL_REGEX.match(email) is not None
    
    @staticmethod
    def validate_name(name: str) -> bool:
        """Validate user name"""
        if not name or len(name) > 100:
            return False
        # Only allow letters, spaces, and common punctuation
        return bool(re.match(r"^[a-zA-Z0-9\s\-'\.]+$", name))
    
    @staticmethod
    def sanitize_string(text: str, max_length: int = 500) -> str:
        """Sanitize string input"""
        if not text:
            return ""
        # Remove null bytes and control characters
        text = text.replace('\x00', '').replace('\n', ' ').replace('\r', '')
        # Limit length
        return text[:max_length].strip()
    
    @staticmethod
    def validate_sheet_id(sheet_id: str) -> bool:
        """Validate Google Sheets ID format"""
        # Google Sheets IDs are typically 44 characters of alphanumeric, -, and _
        if not sheet_id or len(sheet_id) > 100:
            return False
        return bool(re.match(r"^[a-zA-Z0-9\-_]+$", sheet_id))
    
    @staticmethod
    def validate_phone(phone: str) -> bool:
        """Validate phone number format"""
        if not phone or len(phone) > 20:
            return False
        # Remove common formatting characters
        cleaned = re.sub(r'[\s\-\(\)\.+]', '', phone)
        # Should be mostly digits
        return len(cleaned) >= 7 and sum(c.isdigit() for c in cleaned) >= 7


class SessionManager:
    """Manage user sessions with timeout"""
    def __init__(self, timeout_seconds: int = 3600):
        self.timeout_seconds = timeout_seconds
        self.sessions: dict[int, tuple[Any, float]] = {}
    
    def get(self, user_id: int) -> Any | None:
        """Get session, return None if expired"""
        if user_id not in self.sessions:
            return None
        
        session, created_time = self.sessions[user_id]
        if time.time() - created_time > self.timeout_seconds:
            logger.warning(f"⏰ Session expired for user {user_id}")
            del self.sessions[user_id]
            return None
        
        return session
    
    def set(self, user_id: int, session: Any) -> None:
        """Set or update session"""
        self.sessions[user_id] = (session, time.time())
    
    def delete(self, user_id: int) -> None:
        """Delete session"""
        self.sessions.pop(user_id, None)
    
    def cleanup_expired(self) -> None:
        """Remove all expired sessions"""
        now = time.time()
        expired = [
            uid for uid, (_, created) in self.sessions.items()
            if now - created > self.timeout_seconds
        ]
        for uid in expired:
            del self.sessions[uid]
        if expired:
            logger.info(f"🧹 Cleaned up {len(expired)} expired sessions")


class FileValidator:
    """Validate uploaded files for security"""
    
    # Maximum file sizes (in bytes)
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
    MAX_EXCEL_SIZE = 5 * 1024 * 1024  # 5 MB
    
    # Allowed file extensions
    ALLOWED_EXTENSIONS = {'.xlsx', '.xls', '.xlsm'}
    
    # Suspicious patterns in Excel files
    SUSPICIOUS_PATTERNS = [
        b'cmd.exe',
        b'powershell',
        b'bash',
        b'eval(',
        b'exec(',
        b'__import__',
    ]
    
    @staticmethod
    def validate_excel_file(file_name: str, file_content: bytes) -> tuple[bool, str]:
        """
        Validate Excel file for security
        Returns: (is_valid, error_message)
        """
        # 1. Check file extension
        _, ext = file_name.rsplit('.', 1) if '.' in file_name else ('', '')
        ext = f".{ext.lower()}"
        
        if ext not in FileValidator.ALLOWED_EXTENSIONS:
            return False, f"❌ Fayl turi ruxsatga tushmasligi kerak: {ext}"
        
        # 2. Check file size
        file_size = len(file_content)
        if file_size > FileValidator.MAX_EXCEL_SIZE:
            size_mb = file_size / (1024 * 1024)
            max_mb = FileValidator.MAX_EXCEL_SIZE / (1024 * 1024)
            return False, f"❌ Fayl juda katta ({size_mb:.1f} MB). Max: {max_mb:.0f} MB"
        
        if file_size == 0:
            return False, "❌ Fayl bo'sh."
        
        # 3. Check for suspicious patterns
        for pattern in FileValidator.SUSPICIOUS_PATTERNS:
            if pattern in file_content:
                logger.warning(f"🚨 Suspicious pattern detected in file: {pattern}")
                return False, "❌ Fayl xavfsiz emas (shubhali kontent topildi)"
        
        # 4. Try to parse the file to ensure it's valid
        try:
            if ext == '.xlsx':
                from openpyxl import load_workbook
                workbook = load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
                
                # Check for VBA macros (potential security risk)
                if hasattr(workbook, 'vba_archive') and workbook.vba_archive:
                    logger.warning(f"⚠️ File contains VBA macros")
                    return False, "❌ Fayl VBA makros o'z ichiga oladi. Iltimos, sof faylni yuboring."
                
                # Get basic info
                sheet_count = len(workbook.sheetnames)
                logger.info(f"✅ Excel file validated: {sheet_count} sheets")
                
            elif ext == '.xls':
                import xlrd
                workbook = xlrd.open_workbook(file_contents=file_content, on_demand=True)
                sheet_count = workbook.nsheets
                logger.info(f"✅ Excel file validated: {sheet_count} sheets")
        
        except Exception as e:
            logger.error(f"❌ Failed to parse Excel file: {e}")
            return False, f"❌ Faylni o'qib bo'lmadi: {str(e)[:50]}"
        
        return True, "✅ Fayl tekshirildi"
    
    @staticmethod
    def sanitize_filename(filename: str) -> str:
        """Sanitize filename to prevent directory traversal"""
        # Remove path separators
        filename = filename.replace('\\', '').replace('/', '')
        # Remove null bytes
        filename = filename.replace('\x00', '')
        # Limit length
        return filename[:255]


# Initialize security components
rate_limiter = RateLimiter(max_requests=20, time_window=60)  # 20 requests per minute
input_validator = InputValidator()

EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# ---------------------------------------------------------------------------
# SQLite-backed Google token store
# Tokens are persisted to a local SQLite file so they survive bot restarts.
# Supabase is still used as a secondary backup; SQLite is the primary fast
# store that works even when Supabase is unavailable.
# ---------------------------------------------------------------------------

class SQLiteTokenStore:
    """Persist Google OAuth tokens in a local SQLite database."""

    # Use SQLITE_TOKEN_DB env var; if the directory doesn't exist, fall back to /tmp
    _raw_path = os.environ.get("SQLITE_TOKEN_DB", "google_tokens.db")
    _db_dir = os.path.dirname(_raw_path) if os.path.dirname(_raw_path) else "."
    DB_PATH = _raw_path if (os.path.exists(_db_dir) or _db_dir == ".") else os.path.join("/tmp", os.path.basename(_raw_path))

    def __init__(self) -> None:
        self._init_db()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.DB_PATH, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self) -> None:
        try:
            with self._connect() as conn:
                conn.execute("""
                    CREATE TABLE IF NOT EXISTS google_tokens (
                        telegram_id   INTEGER PRIMARY KEY,
                        credentials_json TEXT NOT NULL,
                        updated_at    TEXT NOT NULL
                    )
                """)
                conn.commit()
            logger.info("✅ SQLite token store initialised")
        except Exception as exc:
            logger.error(f"❌ SQLite token store init error: {exc}")

    def save(self, telegram_id: int, credentials_json: str) -> None:
        try:
            now = datetime.now(timezone.utc).isoformat()
            with self._connect() as conn:
                conn.execute("""
                    INSERT INTO google_tokens (telegram_id, credentials_json, updated_at)
                    VALUES (?, ?, ?)
                    ON CONFLICT(telegram_id) DO UPDATE SET
                        credentials_json = excluded.credentials_json,
                        updated_at       = excluded.updated_at
                """, (telegram_id, credentials_json, now))
                conn.commit()
            logger.info(f"💾 SQLite: saved token for user {telegram_id}")
        except Exception as exc:
            logger.warning(f"⚠️ SQLite token save error: {exc}")

    def load(self, telegram_id: int) -> str | None:
        try:
            with self._connect() as conn:
                row = conn.execute(
                    "SELECT credentials_json FROM google_tokens WHERE telegram_id = ?",
                    (telegram_id,)
                ).fetchone()
            if row:
                logger.info(f"✅ SQLite: loaded token for user {telegram_id}")
                return row["credentials_json"]
        except Exception as exc:
            logger.warning(f"⚠️ SQLite token load error: {exc}")
        return None

    def delete(self, telegram_id: int) -> None:
        try:
            with self._connect() as conn:
                conn.execute("DELETE FROM google_tokens WHERE telegram_id = ?", (telegram_id,))
                conn.commit()
        except Exception as exc:
            logger.warning(f"⚠️ SQLite token delete error: {exc}")


# Singleton token store – imported everywhere inside this module
_token_store = SQLiteTokenStore()

# ---------------------------------------------------------------------------
# SQLite-backed workspace store
# Persists connected folder/sheet metadata and cached spreadsheet data so
# users never have to re-upload a link after a bot restart.
#
# Schema (future-proof multi-tenant structure):
#   companies       – one company per admin; holds folder_id / folder_url
#   company_users   – links telegram_id → company_id
#   sheets_cache    – cached JSON data per sheet_id
# ---------------------------------------------------------------------------

class WorkspaceStore:
    """Persist folder/sheet connections and cached data across restarts."""

    # Use same fallback logic as SQLiteTokenStore
    _raw_path = os.environ.get("SQLITE_TOKEN_DB", "google_tokens.db")
    _db_dir = os.path.dirname(_raw_path) if os.path.dirname(_raw_path) else "."
    DB_PATH = _raw_path if (os.path.exists(_db_dir) or _db_dir == ".") else os.path.join("/tmp", os.path.basename(_raw_path))

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.DB_PATH, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")  # safe concurrent reads
        return conn

    def init_db(self) -> None:
        try:
            with self._connect() as conn:
                conn.executescript("""
                    CREATE TABLE IF NOT EXISTS companies (
                        company_id   INTEGER PRIMARY KEY AUTOINCREMENT,
                        admin_user_id INTEGER NOT NULL UNIQUE,
                        folder_id    TEXT,
                        folder_url   TEXT,
                        sheet_id     TEXT,
                        sheet_name   TEXT,
                        mode         TEXT DEFAULT 'sheets',
                        folder_spreadsheets TEXT,
                        selected_spreadsheets TEXT,
                        connected_at TEXT NOT NULL
                    );

                    CREATE TABLE IF NOT EXISTS company_users (
                        user_id    INTEGER PRIMARY KEY,
                        company_id INTEGER NOT NULL,
                        joined_at  TEXT NOT NULL,
                        FOREIGN KEY (company_id) REFERENCES companies(company_id)
                    );

                    CREATE TABLE IF NOT EXISTS sheets_cache (
                        cache_key    TEXT PRIMARY KEY,
                        company_id   INTEGER NOT NULL,
                        sheet_name   TEXT,
                        data         TEXT NOT NULL,
                        last_updated TEXT NOT NULL,
                        FOREIGN KEY (company_id) REFERENCES companies(company_id)
                    );
                """)
                conn.commit()
            logger.info("✅ SQLite workspace store initialised")
        except Exception as exc:
            logger.error(f"❌ SQLite workspace store init error: {exc}")

    # ------------------------------------------------------------------
    # Save / load a user's workspace (folder or single-sheet connection)
    # ------------------------------------------------------------------

    def save_workspace(
        self,
        telegram_id: int,
        *,
        mode: str,                              # "folder" | "sheets"
        folder_id: str | None = None,
        folder_url: str | None = None,
        sheet_id: str | None = None,
        sheet_name: str | None = None,
        folder_spreadsheets: list | None = None,
        selected_spreadsheets: list | None = None,
    ) -> int:
        """Upsert a company row for this admin and return company_id.
        Saves to SQLite first (fast), then mirrors to Supabase (persistent)."""
        now = datetime.now(timezone.utc).isoformat()
        fs_json  = json.dumps(folder_spreadsheets or [])
        sel_json = json.dumps(selected_spreadsheets or [])
        company_id = -1

        # 1. SQLite
        try:
            with self._connect() as conn:
                conn.execute("""
                    INSERT INTO companies
                        (admin_user_id, folder_id, folder_url, sheet_id, sheet_name,
                         mode, folder_spreadsheets, selected_spreadsheets, connected_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(admin_user_id) DO UPDATE SET
                        folder_id              = excluded.folder_id,
                        folder_url             = excluded.folder_url,
                        sheet_id               = excluded.sheet_id,
                        sheet_name             = excluded.sheet_name,
                        mode                   = excluded.mode,
                        folder_spreadsheets    = excluded.folder_spreadsheets,
                        selected_spreadsheets  = excluded.selected_spreadsheets,
                        connected_at           = excluded.connected_at
                """, (telegram_id, folder_id, folder_url, sheet_id, sheet_name,
                      mode, fs_json, sel_json, now))
                row = conn.execute(
                    "SELECT company_id FROM companies WHERE admin_user_id = ?",
                    (telegram_id,)
                ).fetchone()
                company_id = row["company_id"]
                conn.execute("""
                    INSERT INTO company_users (user_id, company_id, joined_at)
                    VALUES (?, ?, ?)
                    ON CONFLICT(user_id) DO UPDATE SET
                        company_id = excluded.company_id,
                        joined_at  = excluded.joined_at
                """, (telegram_id, company_id, now))
                conn.commit()
            logger.info(f"💾 SQLite workspace saved for user {telegram_id} (company {company_id})")
        except Exception as exc:
            logger.warning(f"⚠️ SQLite save_workspace error: {exc}")

        # 2. Supabase mirror (belt-and-suspenders, survives container wipes)
        try:
            from supabase import create_client as _sc
            _sb = _sc(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_ANON_KEY"))
            payload = {
                "admin_user_id": telegram_id,
                "folder_id": folder_id,
                "folder_url": folder_url,
                "sheet_id": sheet_id,
                "sheet_name": sheet_name,
                "mode": mode,
                "folder_spreadsheets": fs_json,
                "selected_spreadsheets": sel_json,
                "connected_at": now,
            }
            resp = _sb.table("workspaces").upsert(payload, on_conflict="admin_user_id").execute()
            if resp.data:
                sb_cid = resp.data[0].get("company_id") or resp.data[0].get("id")
                logger.info(f"💾 Supabase workspace saved for user {telegram_id}")
                # If SQLite didn't assign a company_id, use Supabase's
                if company_id < 0 and sb_cid:
                    company_id = int(sb_cid)
        except Exception as exc:
            logger.warning(f"⚠️ Supabase save_workspace error (SQLite backup ok): {exc}")

        return company_id

    def load_workspace(self, telegram_id: int) -> dict | None:
        """Return the saved workspace dict for this user, or None.
        Tries SQLite first, falls back to Supabase, backfills SQLite."""
        # 1. SQLite
        try:
            with self._connect() as conn:
                row = conn.execute("""
                    SELECT c.*
                    FROM companies c
                    JOIN company_users cu ON cu.company_id = c.company_id
                    WHERE cu.user_id = ?
                    ORDER BY c.connected_at DESC
                    LIMIT 1
                """, (telegram_id,)).fetchone()
            if not row:
                with self._connect() as conn:
                    row = conn.execute(
                        "SELECT * FROM companies WHERE admin_user_id = ? ORDER BY connected_at DESC LIMIT 1",
                        (telegram_id,)
                    ).fetchone()
            if row:
                return dict(row)
        except Exception as exc:
            logger.warning(f"⚠️ SQLite load_workspace error: {exc}")

        # 2. Supabase fallback
        try:
            from supabase import create_client as _sc
            _sb = _sc(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_ANON_KEY"))
            resp = _sb.table("workspaces").select("*").eq("admin_user_id", telegram_id).order("connected_at", desc=True).limit(1).execute()
            if resp.data:
                ws = resp.data[0]
                logger.info(f"✅ Supabase→SQLite workspace backfill for user {telegram_id}")
                # Backfill SQLite
                try:
                    now = datetime.now(timezone.utc).isoformat()
                    with self._connect() as conn:
                        conn.execute("""
                            INSERT INTO companies
                                (admin_user_id, folder_id, folder_url, sheet_id, sheet_name,
                                 mode, folder_spreadsheets, selected_spreadsheets, connected_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ON CONFLICT(admin_user_id) DO UPDATE SET
                                folder_id=excluded.folder_id, folder_url=excluded.folder_url,
                                sheet_id=excluded.sheet_id, sheet_name=excluded.sheet_name,
                                mode=excluded.mode, folder_spreadsheets=excluded.folder_spreadsheets,
                                selected_spreadsheets=excluded.selected_spreadsheets,
                                connected_at=excluded.connected_at
                        """, (telegram_id, ws.get("folder_id"), ws.get("folder_url"),
                              ws.get("sheet_id"), ws.get("sheet_name"), ws.get("mode", "sheets"),
                              ws.get("folder_spreadsheets", "[]"), ws.get("selected_spreadsheets", "[]"),
                              ws.get("connected_at", now)))
                        sqlite_row = conn.execute(
                            "SELECT company_id FROM companies WHERE admin_user_id = ?", (telegram_id,)
                        ).fetchone()
                        if sqlite_row:
                            conn.execute("""
                                INSERT INTO company_users (user_id, company_id, joined_at)
                                VALUES (?, ?, ?)
                                ON CONFLICT(user_id) DO UPDATE SET company_id=excluded.company_id
                            """, (telegram_id, sqlite_row["company_id"], now))
                        conn.commit()
                except Exception as bf_exc:
                    logger.debug(f"SQLite backfill error: {bf_exc}")
                return ws
        except Exception as exc:
            logger.warning(f"⚠️ Supabase load_workspace error: {exc}")

        return None

    def delete_workspace(self, telegram_id: int) -> None:
        """Remove workspace + cache for this user (used by /disconnect)."""
        # 1. SQLite
        try:
            with self._connect() as conn:
                row = conn.execute(
                    "SELECT company_id FROM companies WHERE admin_user_id = ?",
                    (telegram_id,)
                ).fetchone()
                if row:
                    company_id = row["company_id"]
                    conn.execute("DELETE FROM sheets_cache WHERE company_id = ?", (company_id,))
                    conn.execute("DELETE FROM company_users WHERE company_id = ?", (company_id,))
                    conn.execute("DELETE FROM companies WHERE company_id = ?", (company_id,))
                else:
                    conn.execute("DELETE FROM company_users WHERE user_id = ?", (telegram_id,))
                conn.commit()
            logger.info(f"🗑️ SQLite workspace deleted for user {telegram_id}")
        except Exception as exc:
            logger.warning(f"⚠️ SQLite delete_workspace error: {exc}")

        # 2. Supabase
        try:
            from supabase import create_client as _sc
            _sb = _sc(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_ANON_KEY"))
            _sb.table("workspaces").delete().eq("admin_user_id", telegram_id).execute()
            _sb.table("sheets_cache_sb").delete().eq("admin_user_id", telegram_id).execute()
            logger.info(f"🗑️ Supabase workspace deleted for user {telegram_id}")
        except Exception as exc:
            logger.warning(f"⚠️ Supabase delete_workspace error: {exc}")

    # ------------------------------------------------------------------
    # Sheet data cache
    # ------------------------------------------------------------------

    def save_cache(self, company_id: int, cache_key: str, sheet_name: str, data: Any, telegram_id: int = 0) -> None:
        """Save spreadsheet data as JSON to SQLite + Supabase."""
        now = datetime.now(timezone.utc).isoformat()
        try:
            data_json = json.dumps(data, ensure_ascii=False)
        except Exception:
            data_json = "[]"

        # 1. SQLite
        try:
            with self._connect() as conn:
                conn.execute("""
                    INSERT INTO sheets_cache (cache_key, company_id, sheet_name, data, last_updated)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(cache_key) DO UPDATE SET
                        data         = excluded.data,
                        sheet_name   = excluded.sheet_name,
                        last_updated = excluded.last_updated
                """, (cache_key, company_id, sheet_name, data_json, now))
                conn.commit()
            logger.debug(f"💾 SQLite cache saved: {cache_key}")
        except Exception as exc:
            logger.warning(f"⚠️ SQLite save_cache error: {exc}")

        # 2. Supabase (only save if we have a telegram_id to use as owner key)
        if telegram_id:
            try:
                from supabase import create_client as _sc
                _sb = _sc(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_ANON_KEY"))
                _sb.table("sheets_cache_sb").upsert({
                    "cache_key": cache_key,
                    "admin_user_id": telegram_id,
                    "sheet_name": sheet_name,
                    "data": data_json,
                    "last_updated": now,
                }, on_conflict="cache_key").execute()
                logger.debug(f"💾 Supabase cache saved: {cache_key}")
            except Exception as exc:
                logger.debug(f"⚠️ Supabase save_cache error (SQLite ok): {exc}")

    def load_all_cache(self, company_id: int, telegram_id: int = 0) -> dict[str, Any]:
        """Load all cached sheet data. Tries SQLite first, falls back to Supabase."""
        result: dict[str, Any] = {}

        # 1. SQLite
        try:
            with self._connect() as conn:
                rows = conn.execute(
                    "SELECT cache_key, data FROM sheets_cache WHERE company_id = ?",
                    (company_id,)
                ).fetchall()
            for row in rows:
                try:
                    result[row["cache_key"]] = json.loads(row["data"])
                except Exception:
                    pass
        except Exception as exc:
            logger.warning(f"⚠️ SQLite load_all_cache error: {exc}")

        if result:
            return result

        # 2. Supabase fallback
        if telegram_id:
            try:
                from supabase import create_client as _sc
                _sb = _sc(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_ANON_KEY"))
                resp = _sb.table("sheets_cache_sb").select("cache_key, data, sheet_name").eq("admin_user_id", telegram_id).execute()
                if resp.data:
                    logger.info(f"✅ Supabase→SQLite cache backfill for user {telegram_id}: {len(resp.data)} entries")
                    for entry in resp.data:
                        key = entry["cache_key"]
                        try:
                            parsed = json.loads(entry["data"])
                            result[key] = parsed
                            # Backfill SQLite
                            try:
                                with self._connect() as conn:
                                    conn.execute("""
                                        INSERT INTO sheets_cache (cache_key, company_id, sheet_name, data, last_updated)
                                        VALUES (?, ?, ?, ?, ?)
                                        ON CONFLICT(cache_key) DO UPDATE SET
                                            data=excluded.data, sheet_name=excluded.sheet_name,
                                            last_updated=excluded.last_updated
                                    """, (key, company_id, entry.get("sheet_name", ""), entry["data"],
                                          datetime.now(timezone.utc).isoformat()))
                                    conn.commit()
                            except Exception:
                                pass
                        except Exception:
                            pass
            except Exception as exc:
                logger.warning(f"⚠️ Supabase load_all_cache error: {exc}")

        return result


# Singleton workspace store
_workspace_store = WorkspaceStore()
_workspace_store.init_db()


def save_google_token(telegram_id: int, credentials_json: str) -> None:
    """Save Google credentials to BOTH SQLite (primary) and Supabase (backup)."""
    # 1. SQLite – fast, always available
    _token_store.save(telegram_id, credentials_json)
    # 2. Supabase – persistent across container re-creations
    try:
        from supabase import create_client as _sc
        _sb = _sc(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_ANON_KEY"))
        resp = _sb.table("users").select("telegram_id").eq("telegram_id", telegram_id).execute()
        if resp.data:
            _sb.table("users").update({
                "google_credentials": credentials_json,
                "google_credentials_updated_at": datetime.now(timezone.utc).isoformat(),
            }).eq("telegram_id", telegram_id).execute()
            logger.info(f"💾 Supabase: saved token for user {telegram_id}")
        else:
            logger.warning(f"⚠️ Supabase: user {telegram_id} not found, token not saved to Supabase")
    except Exception as exc:
        logger.warning(f"⚠️ Supabase token save error (SQLite backup still worked): {exc}")


def load_google_token(telegram_id: int) -> str | None:
    """Load Google credentials – tries SQLite first, then Supabase."""
    creds_json = _token_store.load(telegram_id)
    if creds_json:
        return creds_json
    # Fall back to Supabase
    try:
        from supabase import create_client as _sc
        _sb = _sc(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_ANON_KEY"))
        resp = _sb.table("users").select("google_credentials").eq("telegram_id", telegram_id).execute()
        if resp.data and resp.data[0].get("google_credentials"):
            creds_json = resp.data[0]["google_credentials"]
            # Backfill SQLite so next call is fast
            _token_store.save(telegram_id, creds_json)
            logger.info(f"✅ Supabase→SQLite backfill for user {telegram_id}")
            return creds_json
    except Exception as exc:
        logger.warning(f"⚠️ Supabase token load error: {exc}")
    return None


def load_and_refresh_google_token(telegram_id: int) -> str | None:
    """Load token, auto-refresh if expired, save refreshed token, return JSON or None.

    Returns ``None`` only when:
      - No token exists anywhere (user has never authenticated), OR
      - The refresh_token itself is expired/revoked (user must re-authenticate).
    """
    creds_json = load_google_token(telegram_id)
    if not creds_json:
        return None
    try:
        data = json.loads(creds_json)
        creds = Credentials.from_authorized_user_info(data, scopes=SCOPES)

        # Check for old read-only scopes → force re-auth
        stored_scopes = data.get("scopes", [])
        if any("readonly" in s for s in stored_scopes):
            logger.warning(f"⚠️ Old read-only scopes for user {telegram_id}. Re-auth required.")
            return None

        if not creds.valid:
            if creds.expired and creds.refresh_token:
                logger.info(f"🔄 Token expired for user {telegram_id}, refreshing...")
                creds.refresh(GoogleAuthRequest())
                refreshed_json = creds.to_json()
                # Persist the freshly-issued token immediately
                save_google_token(telegram_id, refreshed_json)
                logger.info(f"✅ Token refreshed and saved for user {telegram_id}")
                return refreshed_json
            else:
                # No refresh_token or some other invalidity
                logger.warning(f"⚠️ Token invalid and cannot be refreshed for user {telegram_id}")
                return None
        return creds_json
    except Exception as exc:
        logger.warning(f"⚠️ load_and_refresh_google_token error for user {telegram_id}: {exc}")
        return None


SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",  # Full read/write access to Sheets
    "https://www.googleapis.com/auth/drive",         # Full read/write access to Drive (folders, files, etc)
]
# REDIRECT_URI is now set from config - see Config class
MAIN_MENU_SHEETS = "📊 Google Sheets ulash"
MAIN_MENU_EXCEL = "📁 Excel fayl yuklash"
MAX_ROWS_FOR_CONTEXT = 1000
MAX_COLS_FOR_CONTEXT = 50
MAX_CHARS_CONTEXT = 120000


@dataclass
class Config:
    bot_token: str
    tavily_api_key: str
    # Google OAuth (for Google Sheets access)
    google_client_id: str
    google_client_secret: str
    # GitHub OAuth (for user authentication)
    github_client_id: str
    github_client_secret: str
    # Database
    supabase_url: str
    supabase_anon_key: str
    # Grok AI (xAI) - for intelligent spreadsheet Q&A
    grok_api_key: str = ""
    # Server configuration - can be overridden via env vars
    server_host: str = "0.0.0.0"  # Listen on all interfaces for production
    server_port: int = 8080  # Koyeb / Render both use 8080 by default
    google_redirect_uri: str = "https://onbrain.koyeb.app/"  # For Google OAuth
    github_redirect_uri: str = "https://onbrain.koyeb.app/github/callback"  # For GitHub OAuth

    @classmethod
    def from_env(cls) -> "Config":
        load_dotenv()
        required = {
            "BOT_TOKEN": os.getenv("BOT_TOKEN", "").strip(),
            "TAVILY_API_KEY": os.getenv("TAVILY_API_KEY", "").strip(),
            "GOOGLE_CLIENT_ID": os.getenv("GOOGLE_CLIENT_ID", "").strip(),
            "GOOGLE_CLIENT_SECRET": os.getenv("GOOGLE_CLIENT_SECRET", "").strip(),
            "GITHUB_CLIENT_ID": os.getenv("GITHUB_CLIENT_ID", "").strip(),
            "GITHUB_CLIENT_SECRET": os.getenv("GITHUB_CLIENT_SECRET", "").strip(),
            "SUPABASE_URL": os.getenv("SUPABASE_URL", "").strip(),
            "SUPABASE_ANON_KEY": os.getenv("SUPABASE_ANON_KEY", "").strip(),
        }
        missing = [k for k, v in required.items() if not v]
        if missing:
            raise RuntimeError(
                "Quyidagi .env qiymatlari to'ldirilmagan: " + ", ".join(missing)
            )
        
        # Optional server configuration
        server_host = os.getenv("SERVER_HOST", "0.0.0.0").strip()
        # Koyeb and Render both inject PORT env var; fallback to 8080
        server_port = int(os.getenv("PORT", os.getenv("SERVER_PORT", "8080")))
        
        # Google OAuth redirect URI
        google_redirect_uri = os.getenv("GOOGLE_REDIRECT_URI", "").strip()
        if not google_redirect_uri:
            domain = os.getenv("APP_DOMAIN", "onbrain.koyeb.app").strip()
            if domain == "localhost":
                google_redirect_uri = f"http://localhost:{server_port}/"
            else:
                google_redirect_uri = f"https://{domain}/"
        
        # GitHub OAuth redirect URI
        github_redirect_uri = os.getenv("GITHUB_REDIRECT_URI", "").strip()
        if not github_redirect_uri:
            domain = os.getenv("APP_DOMAIN", "onbrain.koyeb.app").strip()
            if domain == "localhost":
                github_redirect_uri = f"http://localhost:{server_port}/github/callback"
            else:
                github_redirect_uri = f"https://{domain}/github/callback"
        
        # Optional: Grok AI key for spreadsheet Q&A
        grok_api_key = os.getenv("GROK_API_KEY", "").strip()
        
        return cls(
            bot_token=required["BOT_TOKEN"],
            tavily_api_key=required["TAVILY_API_KEY"],
            google_client_id=required["GOOGLE_CLIENT_ID"],
            google_client_secret=required["GOOGLE_CLIENT_SECRET"],
            github_client_id=required["GITHUB_CLIENT_ID"],
            github_client_secret=required["GITHUB_CLIENT_SECRET"],
            supabase_url=required["SUPABASE_URL"],
            supabase_anon_key=required["SUPABASE_ANON_KEY"],
            grok_api_key=grok_api_key,
            server_host=server_host,
            server_port=server_port,
            google_redirect_uri=google_redirect_uri,
            github_redirect_uri=github_redirect_uri,
        )


@dataclass
class UserSession:
    step: str = "idle"
    full_name: str | None = None
    email: str | None = None
    phone_number: str | None = None
    sheet_id: str | None = None
    sheet_name: str | None = None
    sheet_data: list[list[Any]] = field(default_factory=list)
    excel_data: list[list[Any]] = field(default_factory=list)
    all_sheets_data: dict[str, list[list[Any]]] = field(default_factory=dict)  # All sheets from Google Sheets
    google_credentials_json: str | None = None
    pending_sheets: dict[str, str] = field(default_factory=dict)
    # For folder mode
    folder_spreadsheets: list[dict[str, str]] = field(default_factory=list)  # List of {id, name, url}
    selected_spreadsheets: list[str] = field(default_factory=list)  # Selected sheet IDs
    all_folder_sheets_data: dict[str, dict[str, list[list[Any]]]] = field(default_factory=dict)  # sheet_id -> {sheet_name -> data}
    auth_mode: str | None = None  # Track which button was clicked: "sheets" or "folder"
    # GitHub OAuth
    github_username: str | None = None
    github_email: str | None = None
    github_access_token: str | None = None
    # Data Indexing
    indexing_service: Any = None  # DataIndexingService instance
    folder_id: str | None = None  # Current folder ID being indexed
    web_search_mode: bool = False  # When True, bypass spreadsheet and use Tavily


class SessionStore:
    def __init__(self, timeout_seconds: int = 86400) -> None:  # 24 hours
        self._store: dict[int, UserSession] = {}
        self._timestamps: dict[int, float] = {}
        self.timeout_seconds = timeout_seconds

    def get(self, telegram_id: int) -> UserSession:
        # Check if session exists and is not expired
        if telegram_id in self._timestamps:
            if time.time() - self._timestamps[telegram_id] > self.timeout_seconds:
                logger.warning(f"⏰ Session expired for user {telegram_id}")
                self._store.pop(telegram_id, None)
                self._timestamps.pop(telegram_id, None)
        
        if telegram_id not in self._store:
            self._store[telegram_id] = UserSession()
            self._timestamps[telegram_id] = time.time()
        
        # ---------------------------------------------------------------
        # Load Google credentials if not already in memory.
        # ---------------------------------------------------------------
        if not self._store[telegram_id].google_credentials_json:
            refreshed = load_and_refresh_google_token(telegram_id)
            if refreshed:
                self._store[telegram_id].google_credentials_json = refreshed
                logger.info(f"✅ Loaded (and refreshed if needed) Google token for user {telegram_id}")
            # Also load GitHub credentials from Supabase
            try:
                from supabase import create_client
                supabase = create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_ANON_KEY"))
                user_response = supabase.table("users").select("github_username, github_email, github_access_token").eq("telegram_id", telegram_id).execute()
                if user_response.data:
                    if user_response.data[0].get("github_username"):
                        self._store[telegram_id].github_username = user_response.data[0]["github_username"]
                        self._store[telegram_id].github_email = user_response.data[0].get("github_email")
                        self._store[telegram_id].github_access_token = user_response.data[0].get("github_access_token")
                        logger.info(f"✅ Loaded GitHub credentials from database for user {telegram_id}")
            except Exception as e:
                logger.debug(f"⚠️ Could not load GitHub credentials from database: {e}")

            # ---------------------------------------------------------------
            # Restore workspace (folder / sheet connection) from DB.
            # ---------------------------------------------------------------
            self._restore_workspace(telegram_id)

        else:
            # Update timestamp on access
            self._timestamps[telegram_id] = time.time()
            # Also restore workspace if data is missing (e.g. after partial session)
            sess = self._store[telegram_id]
            if not sess.all_folder_sheets_data and not sess.all_sheets_data:
                self._restore_workspace(telegram_id)
        
        return self._store[telegram_id]

    def _restore_workspace(self, telegram_id: int) -> None:
        """Silently restore folder/sheet connection from SQLite workspace store."""
        sess = self._store[telegram_id]
        # Skip if data is already in memory
        if sess.all_folder_sheets_data or sess.all_sheets_data:
            return
        try:
            ws = _workspace_store.load_workspace(telegram_id)
            if not ws:
                return
            company_id = ws["company_id"]
            mode = ws.get("mode", "sheets")
            sess.folder_id = ws.get("folder_id")
            sess.sheet_id  = ws.get("sheet_id")
            sess.sheet_name = ws.get("sheet_name")
            # Restore spreadsheet list metadata
            if ws.get("folder_spreadsheets"):
                try:
                    sess.folder_spreadsheets = json.loads(ws["folder_spreadsheets"])
                except Exception:
                    pass
            if ws.get("selected_spreadsheets"):
                try:
                    sess.selected_spreadsheets = json.loads(ws["selected_spreadsheets"])
                except Exception:
                    pass
            # Restore cached sheet data
            cached = _workspace_store.load_all_cache(company_id, telegram_id=telegram_id)
            if mode == "folder" and cached:
                # Rebuild all_folder_sheets_data: {sheet_id: {sheet_title: rows}}
                folder_data: dict[str, dict[str, list]] = {}
                for cache_key, data in cached.items():
                    # cache_key format: "folder:{sheet_id}:{sheet_title}"
                    if cache_key.startswith("folder:"):
                        parts = cache_key.split(":", 2)
                        if len(parts) == 3:
                            _, sid, stitle = parts
                            if sid not in folder_data:
                                folder_data[sid] = {}
                            folder_data[sid][stitle] = data
                if folder_data:
                    sess.all_folder_sheets_data = folder_data
                    sess.step = "in_chat"
                    logger.info(f"✅ Restored folder workspace for user {telegram_id}: {len(folder_data)} spreadsheets")
            elif mode == "sheets" and cached:
                # Rebuild all_sheets_data: {sheet_title: rows}
                sheets_data: dict[str, list] = {}
                for cache_key, data in cached.items():
                    if cache_key.startswith("sheet:"):
                        parts = cache_key.split(":", 2)
                        if len(parts) == 3:
                            _, _, stitle = parts
                            sheets_data[stitle] = data
                if sheets_data:
                    sess.all_sheets_data = sheets_data
                    sess.step = "in_chat"
                    logger.info(f"✅ Restored sheet workspace for user {telegram_id}: {len(sheets_data)} tabs")
        except Exception as exc:
            logger.warning(f"⚠️ _restore_workspace error for user {telegram_id}: {exc}")
    
    def cleanup_expired(self) -> None:
        """Remove expired sessions"""
        now = time.time()
        expired = [
            uid for uid, ts in self._timestamps.items()
            if now - ts > self.timeout_seconds
        ]
        for uid in expired:
            self._store.pop(uid, None)
            self._timestamps.pop(uid, None)
        if expired:
            logger.info(f"🧹 Cleaned up {len(expired)} expired sessions")


class SupabaseService:
    def __init__(self, supabase_url: str, supabase_key: str) -> None:
        self.url = supabase_url.rstrip('/')
        self.key = supabase_key
        self.headers = {
            "apikey": supabase_key,
            "Authorization": f"Bearer {supabase_key}",
            "Content-Type": "application/json",
        }

    async def get_user_by_telegram(self, telegram_id: int) -> dict[str, Any] | None:
        return await asyncio.to_thread(self._get_user_by_telegram_sync, telegram_id)

    def _get_user_by_telegram_sync(self, telegram_id: int) -> dict[str, Any] | None:
        try:
            response = httpx.get(
                f"{self.url}/rest/v1/users?telegram_id=eq.{telegram_id}&limit=1",
                headers=self.headers,
            )
            data = response.json()
            return data[0] if data else None
        except Exception as exc:
            logger.error(f"Get user by telegram error: {exc}")
            return None

    async def get_user_by_email(self, email: str) -> dict[str, Any] | None:
        return await asyncio.to_thread(self._get_user_by_email_sync, email)

    def _get_user_by_email_sync(self, email: str) -> dict[str, Any] | None:
        try:
            response = httpx.get(
                f"{self.url}/rest/v1/users?email=eq.{email}&limit=1",
                headers=self.headers,
            )
            data = response.json()
            return data[0] if data else None
        except Exception as exc:
            logger.error(f"Get user by email error: {exc}")
            return None

    async def create_user(
        self, 
        telegram_id: int, 
        full_name: str, 
        email: str,
        phone_number: str | None = None,
    ) -> bool:
        """Create user - returns True if successful"""
        try:
            result = await asyncio.to_thread(
                self._create_user_sync, 
                telegram_id, 
                full_name, 
                email,
                phone_number,
            )
            return result
        except Exception as exc:
            logger.error(f"create_user async wrapper error: {exc}")
            return False

    def _create_user_sync(
        self, 
        telegram_id: int, 
        full_name: str, 
        email: str,
        phone_number: str | None = None,
    ) -> bool:
        """Create user in Supabase - returns True if successful"""
        try:
            payload = {
                "telegram_id": telegram_id,
                "first_name": full_name.split()[0] if full_name else "User",
                "last_name": " ".join(full_name.split()[1:]) if len(full_name.split()) > 1 else "",
                "email": email,
            }
            # Add phone_number if provided
            if phone_number:
                payload["phone_number"] = phone_number
            
            logger.info(f"📤 Creating user {telegram_id} with payload: {payload}")
            
            response = httpx.post(
                f"{self.url}/rest/v1/users",
                headers=self.headers,
                json=payload,
            )
            
            logger.info(f"📊 Supabase response status: {response.status_code}")
            
            # Check for status code errors
            if response.status_code >= 400:
                error_detail = response.text
                logger.error(f"Supabase error {response.status_code}: {error_detail}")
                
                # If phone_number is causing issues, retry without it
                if "phone_number" in error_detail.lower() and phone_number:
                    logger.warning(f"⚠️  Phone number field error, retrying without phone_number...")
                    payload.pop("phone_number", None)
                    response = httpx.post(
                        f"{self.url}/rest/v1/users",
                        headers=self.headers,
                        json=payload,
                    )
                    if response.status_code < 400:
                        logger.info(f"✅ User {telegram_id} created successfully (without phone)")
                        return True
                
                raise Exception(f"Supabase error: {error_detail}")
            
            response.raise_for_status()
            logger.info(f"✅ User {telegram_id} created successfully")
            return True
            
        except Exception as exc:
            logger.error(f"❌ Create user error: {exc}")
            return False

    async def save_integration(self, telegram_id: int, sheet_id: str, sheet_name: str) -> None:
        await asyncio.to_thread(self._save_integration_sync, telegram_id, sheet_id, sheet_name)

    def _save_integration_sync(self, telegram_id: int, sheet_id: str, sheet_name: str) -> None:
        try:
            # Deactivate old integrations
            httpx.patch(
                f"{self.url}/rest/v1/integrations?telegram_id=eq.{telegram_id}",
                headers=self.headers,
                json={"is_active": False},
            )
            # Create new integration
            payload = {
                "telegram_id": telegram_id,
                "sheet_id": sheet_id,
                "sheet_name": sheet_name,
                "is_active": True,
            }
            response = httpx.post(
                f"{self.url}/rest/v1/integrations",
                headers=self.headers,
                json=payload,
            )
            response.raise_for_status()
        except Exception as exc:
            logger.error(f"Save integration error: {exc}")
            raise

    async def get_active_integration(self, telegram_id: int) -> dict[str, Any] | None:
        return await asyncio.to_thread(self._get_active_integration_sync, telegram_id)

    def _get_active_integration_sync(self, telegram_id: int) -> dict[str, Any] | None:
        try:
            response = httpx.get(
                f"{self.url}/rest/v1/integrations?telegram_id=eq.{telegram_id}&is_active=eq.true&order=created_at.desc&limit=1",
                headers=self.headers,
            )
            data = response.json()
            return data[0] if data else None
        except Exception as exc:
            logger.error(f"Get active integration error: {exc}")
            return None

    async def save_message(self, telegram_id: int, question: str, answer: str) -> None:
        await asyncio.to_thread(self._save_message_sync, telegram_id, question, answer)

    def _save_message_sync(self, telegram_id: int, question: str, answer: str) -> None:
        payload = {
            "id": str(uuid.uuid4()),
            "telegram_id": telegram_id,
            "question": question,
            "answer": answer,
        }
        self.client.table("messages").insert(payload).execute()


class GoogleOAuthService:
    def __init__(self, client_id: str, client_secret: str, redirect_uri: str = "https://onbrain.koyeb.app/") -> None:
        self.client_config = {
            "web": {
                "client_id": client_id,
                "client_secret": client_secret,
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
            }
        }
        self.redirect_uri = redirect_uri
        self.pending_flows: dict[str, dict[str, Any]] = {}

    def create_auth_url(self, telegram_id: int) -> str:
        flow = Flow.from_client_config(
            self.client_config, scopes=SCOPES, redirect_uri=self.redirect_uri
        )
        auth_url, state = flow.authorization_url(
            access_type="offline",
            include_granted_scopes="true",
            prompt="consent",
        )
        self.pending_flows[state] = {
            "flow": flow,
            "telegram_id": telegram_id,
            "created_at": time.time(),
        }
        logger.info(f"🔑 Created OAuth state for user {telegram_id}. Total pending: {len(self.pending_flows)}. State prefix: {state[:12]}...")
        self.cleanup_stale_flows()
        return auth_url

    def cleanup_stale_flows(self) -> None:
        now = time.time()
        # Increase timeout to 1 hour (3600 seconds) instead of 15 minutes
        # This prevents "OAuth state expired" errors if user takes time to authorize
        stale = [state for state, item in self.pending_flows.items() if now - item["created_at"] > 3600]
        for state in stale:
            self.pending_flows.pop(state, None)
            logger.debug(f"Cleaned up stale OAuth state (older than 1 hour)")

    def exchange_code(self, state: str, code: str) -> tuple[int, Credentials]:
        logger.info(f"🔑 Exchange code called. State prefix: {state[:12]}... Total pending: {len(self.pending_flows)}")
        if state not in self.pending_flows:
            logger.warning(f"❌ OAuth state not found. State prefix: {state[:12]}... Available states: {len(self.pending_flows)}")
            # Log available state prefixes for debugging
            for s in self.pending_flows:
                logger.warning(f"   Available state prefix: {s[:12]}...")
            raise ValueError("OAuth holati topilmadi yoki muddati tugagan. Iltimos, qayta urinib ko'ring. (State not in pending flows)")
        flow_item = self.pending_flows.pop(state)
        flow: Flow = flow_item["flow"]
        flow.fetch_token(code=code)
        telegram_id: int = flow_item["telegram_id"]
        return telegram_id, flow.credentials


# ---------------------------------------------------------------------------
# Safe HTML message helpers – catch TelegramBadRequest and fall back to plain
# text so the bot never crashes on an unparseable HTML string.
# ---------------------------------------------------------------------------

def _strip_html_tags(text: str) -> str:
    """Remove common HTML tags used in bot messages."""
    import re as _re
    return _re.sub(r"</?(?:b|i|u|s|code|pre|a)[^>]*>", "", text)


def _escape_url_for_html(url: str) -> str:
    """Escape a URL so it is safe inside HTML parse_mode messages.

    Telegram's HTML parser chokes on bare ``&`` (and ``<``, ``>``) inside
    attribute values and message text.  This helper encodes them.
    """
    return url.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _extract_sheet_id(text: str) -> str | None:
    """Extract a Google Sheets spreadsheet ID from *any* common link format.

    Supported patterns:
      • https://docs.google.com/spreadsheets/d/SHEET_ID/edit#gid=0
      • https://docs.google.com/spreadsheets/d/SHEET_ID/edit?usp=sharing
      • https://docs.google.com/spreadsheets/d/SHEET_ID/
      • https://docs.google.com/spreadsheets/d/SHEET_ID
      • https://docs.google.com/spreadsheets/d/SHEET_ID/export?format=csv
      • https://docs.google.com/spreadsheets/d/SHEET_ID/gviz/tq
      • https://docs.google.com/spreadsheets/d/SHEET_ID/pub
      • https://docs.google.com/spreadsheets/d/SHEET_ID/htmlview
      • https://docs.google.com/spreadsheets/d/e/LONG_PUB_ID/pubhtml (published)
      • https://sheets.googleapis.com/v4/spreadsheets/SHEET_ID
      • Plain sheet ID pasted by the user (44-char alphanumeric string)

    Returns the sheet ID string or ``None`` if nothing matched.
    """
    if not text:
        return None

    text = text.strip()

    # 1. Standard docs.google.com/spreadsheets/d/<ID> URLs
    m = re.search(r'docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]{10,})', text)
    if m:
        return m.group(1)

    # 2. Published "d/e/<pubid>" links — strip the "e/" prefix, use the long ID
    m = re.search(r'docs\.google\.com/spreadsheets/d/e/([a-zA-Z0-9_-]{10,})', text)
    if m:
        return m.group(1)

    # 3. Sheets API URL  sheets.googleapis.com/v4/spreadsheets/<ID>
    m = re.search(r'sheets\.googleapis\.com/v4/spreadsheets/([a-zA-Z0-9_-]{10,})', text)
    if m:
        return m.group(1)

    # 4. User pasted a raw sheet ID (no URL around it)
    m = re.match(r'^([a-zA-Z0-9_-]{20,})$', text)
    if m:
        return m.group(1)

    return None


def _looks_like_sheets_url(text: str) -> bool:
    """Return True if *text* looks like it could be a Google Sheets link,
    even if we can't extract a sheet ID from it.  Used to give the user a
    more helpful error instead of 'Bu Google Sheets linki emas'.
    """
    lower = text.lower()
    return any(kw in lower for kw in (
        "docs.google.com/spreadsheets",
        "sheets.googleapis.com",
        "spreadsheets/d/",
        "sheets.google.com",
        "google.com/spreadsheets",
    ))


async def safe_send(target, text: str, *, parse_mode: str = "HTML", **kwargs):
    """Send *text* via ``target.answer`` (Message) or ``target.edit_text``
    (CallbackQuery.message).  If the Telegram API rejects the HTML, retry
    once with the tags stripped and ``parse_mode`` removed so the user still
    gets a response instead of an error.

    *target* – ``Message`` or the ``.message`` attribute of a ``CallbackQuery``
    """
    send_fn = getattr(target, "answer", None) or getattr(target, "edit_text", None)
    if send_fn is None:
        raise TypeError(f"Cannot send to {type(target)}")
    try:
        return await send_fn(text, parse_mode=parse_mode, **kwargs)
    except (TelegramBadRequest, TelegramAPIError) as exc:
        logger.warning("HTML parse failed, falling back to plain text: %s", exc)
        plain = _strip_html_tags(text)
        try:
            return await send_fn(plain, **kwargs)
        except Exception as inner:
            logger.error("Plain-text fallback also failed: %s", inner)


async def safe_edit(msg, text: str, *, parse_mode: str = "HTML", **kwargs):
    """Like safe_send but specifically for ``callback.message.edit_text``."""
    try:
        return await msg.edit_text(text, parse_mode=parse_mode, **kwargs)
    except (TelegramBadRequest, TelegramAPIError) as exc:
        logger.warning("HTML parse failed on edit, falling back: %s", exc)
        plain = _strip_html_tags(text)
        try:
            return await msg.edit_text(plain, **kwargs)
        except Exception as inner:
            logger.error("Plain-text edit fallback also failed: %s", inner)


def build_main_menu() -> InlineKeyboardMarkup:
    """Build main menu with Chat, Sheets, Excel, Folder, and GitHub buttons"""
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(
                text="💬 Chat with AI",
                callback_data="chat_start"
            )],
            [InlineKeyboardButton(text=MAIN_MENU_SHEETS, callback_data="sheets")],
            [InlineKeyboardButton(text="📁 Google Drive Folder", callback_data="folder")],
            [InlineKeyboardButton(text=MAIN_MENU_EXCEL, callback_data="excel")],
            [InlineKeyboardButton(text="🔐 GitHub'da kirish", callback_data="github_login")],
        ]
    )


def build_assistant_keyboard() -> InlineKeyboardMarkup:
    """Build keyboard with Assistant chat and main menu buttons"""
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🤖 Assistant ga murojaat", callback_data="chat_start")],
            [InlineKeyboardButton(text="🏠 Asosiy menyu ga qaytish", callback_data="main_menu")],
        ]
    )


def build_chat_response_keyboard() -> InlineKeyboardMarkup:
    """Build keyboard for chat responses with continue and exit buttons"""
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="💬 Chat-ga qaytish", callback_data="chat_continue")],
            [InlineKeyboardButton(text="🌐 Internet qidirish", callback_data="web_search_mode")],
            [InlineKeyboardButton(text="🚪 Chat-ni tugatish", callback_data="exit_chat")]
        ]
    )


def build_retry_keyboard(context: str = "sheets") -> InlineKeyboardMarkup:
    """Build keyboard with Retry, Assistant and Main menu buttons
    
    Args:
        context: The context for the retry button ("sheets" or "folder")
    """
    retry_callback = f"retry_{context}"  # "retry_sheets" or "retry_folder"
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🔄 Qayta yuborish", callback_data=retry_callback)],
            [InlineKeyboardButton(text="🤖 Assistant ga murojaat", callback_data="chat_start")],
            [InlineKeyboardButton(text="🏠 Asosiy menyu ga qaytish", callback_data="main_menu")],
        ]
    )


def is_valid_email(email: str) -> bool:
    return bool(EMAIL_REGEX.match(email.strip()))


def limit_2d_table(rows: list[list[Any]]) -> list[list[str]]:
    trimmed: list[list[str]] = []
    for row in rows[:MAX_ROWS_FOR_CONTEXT]:
        trimmed.append([str(col) if col is not None else "" for col in row[:MAX_COLS_FOR_CONTEXT]])
    return trimmed


def table_to_text(rows: list[list[Any]]) -> str:
    trimmed = limit_2d_table(rows)
    if not trimmed:
        return "Ma'lumot topilmadi."
    lines = [" | ".join(row) for row in trimmed]
    text = "\n".join(lines)
    return text[:MAX_CHARS_CONTEXT]


def credentials_from_json(credentials_json: str, telegram_id: int | None = None) -> Credentials:
    """Parse credentials JSON, auto-refresh if expired, and persist the refreshed token.

    Pass ``telegram_id`` so that a freshly-refreshed token is immediately saved
    back to SQLite + Supabase, preventing the need to re-authenticate after
    every bot restart.
    """
    data = json.loads(credentials_json)
    creds = Credentials.from_authorized_user_info(data, scopes=SCOPES)
    
    # Check if stored credentials have old scopes (only .readonly)
    stored_scopes = data.get("scopes", [])
    has_old_scopes = any("readonly" in scope for scope in stored_scopes)
    
    # If old scopes detected, raise error to trigger re-authentication
    if has_old_scopes:
        logger.warning(f"⚠️ Old credential scopes detected (read-only). User needs to re-authenticate.")
        raise ValueError("Credentials have old scopes. Please re-authenticate to grant full access.")
    
    # Refresh if expired, then save the new token so it survives future restarts
    if creds.expired and creds.refresh_token:
        try:
            creds.refresh(GoogleAuthRequest())
            logger.info("✅ Credentials refreshed successfully")
            if telegram_id:
                save_google_token(telegram_id, creds.to_json())
        except Exception as e:
            logger.warning(f"⚠️ Could not refresh credentials: {e}")
    
    return creds


def list_google_sheets(credentials: Credentials) -> list[dict[str, str]]:
    if credentials.expired and credentials.refresh_token:
        credentials.refresh(GoogleAuthRequest())
    gc = gspread.authorize(credentials)
    files = gc.list_spreadsheet_files()
    sheets = [{"id": item["id"], "name": item.get("name", "Nomsiz jadval")} for item in files]
    if sheets:
        return sheets
    books = gc.openall()
    return [{"id": book.id, "name": book.title} for book in books]


def fetch_sheet_rows(credentials: Credentials, sheet_id: str) -> dict[str, list[list[Any]]]:
    """Fetch ALL sheets from a Google Sheets spreadsheet"""
    if credentials.expired and credentials.refresh_token:
        credentials.refresh(GoogleAuthRequest())
    gc = gspread.authorize(credentials)
    workbook = gc.open_by_key(sheet_id)
    
    # Get all worksheets in the spreadsheet
    all_sheets = {}
    for worksheet in workbook.worksheets():
        sheet_name = worksheet.title
        try:
            rows = worksheet.get_all_values()
            all_sheets[sheet_name] = rows
            logger.info(f"📊 Read sheet '{sheet_name}' with {len(rows)} rows")
        except Exception as e:
            logger.warning(f"⚠️ Could not read sheet '{sheet_name}': {e}")
            all_sheets[sheet_name] = []
    
    logger.info(f"✅ Successfully read {len(all_sheets)} sheets from spreadsheet")
    return all_sheets


def parse_excel_bytes(file_name: str, content: bytes) -> list[list[Any]]:
    lower = file_name.lower()
    if lower.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return rows
    if lower.endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)
        rows = []
        for r in range(ws.nrows):
            rows.append(ws.row_values(r))
        return rows
    raise ValueError("Noto'g'ri fayl turi.")


class AppContext:
    def __init__(self, config: Config) -> None:
        self.config = config
        self.supabase_service = SupabaseService(
            config.supabase_url, config.supabase_anon_key
        )
        self.sessions = SessionStore()
        # Google OAuth Service (for Google Sheets)
        self.oauth_service = GoogleOAuthService(
            client_id=config.google_client_id,
            client_secret=config.google_client_secret,
            redirect_uri=config.google_redirect_uri,
        )
        # GitHub OAuth Service (for user authentication)
        self.github_oauth_service = GitHubOAuthService(
            client_id=config.github_client_id,
            client_secret=config.github_client_secret,
            redirect_uri=config.github_redirect_uri,
        )
        self.tavily_api_key = config.tavily_api_key  # Store key for Tavily API
        self.grok_api_key = config.grok_api_key      # Store key for Grok AI (xAI)
        self.bot: Bot | None = None
    
    def _save_credentials_sync(self, telegram_id: int, credentials_json: str) -> None:
        """Synchronously save Google credentials to SQLite + Supabase."""
        save_google_token(telegram_id, credentials_json)
    def _save_github_credentials_sync(self, telegram_id: int, github_username: str, github_email: str, github_token: str) -> None:
        """Synchronously save GitHub credentials to database"""
        try:
            from supabase import create_client
            supabase = create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_ANON_KEY"))
            
            user_response = supabase.table("users").select("*").eq("telegram_id", telegram_id).execute()
            
            if user_response.data:
                supabase.table("users").update({
                    "github_username": github_username,
                    "github_email": github_email,
                    "github_access_token": github_token,
                    "github_authenticated_at": datetime.utcnow().isoformat()
                }).eq("telegram_id", telegram_id).execute()
                logger.info(f"✅ Saved GitHub credentials to database for user {telegram_id}")
            else:
                logger.warning(f"⚠️ User {telegram_id} not found in database for GitHub credential save")
        except Exception as e:
            logger.warning(f"⚠️ Could not save GitHub credentials to database: {e}")


    async def handle_oauth_callback(self, state: str, code: str) -> str:
        try:
            telegram_id, credentials = await asyncio.to_thread(
                self.oauth_service.exchange_code, state, code
            )
            session = self.sessions.get(telegram_id)
            creds_json = credentials.to_json()
            session.google_credentials_json = creds_json
            logger.info(f"✅ OAuth successful for user {telegram_id}")
            
            # Save credentials to SQLite + Supabase immediately so they survive restarts
            await asyncio.to_thread(save_google_token, telegram_id, creds_json)
            
            # Check which mode user chose
            if session.auth_mode == "folder":
                session.step = "waiting_folder_link"
                message_text = "✅ <b>Google hisobiga muvaffaqiyatli ulandi!</b>\n\n" \
                              "Endi Google Drive papka havolasini jo'nating:\n\n" \
                              "📋 <b>Misol:</b>\n" \
                              "<code>https://drive.google.com/drive/folders/1ABC123xyz</code>"
            else:
                # Default to sheets mode
                session.step = "waiting_sheet_link"
                message_text = "✅ <b>Google hisobiga muvaffaqiyatli ulandi!</b>\n\n" \
                              "Endi Google Sheets havolasini jo'nating:\n\n" \
                              "📋 <b>Misol:</b>\n" \
                              "<code>https://docs.google.com/spreadsheets/d/1Abc123xyz/edit</code>"
            
            # Send success message
            if self.bot:
                try:
                    await self.bot.send_message(
                        telegram_id,
                        message_text,
                        parse_mode="HTML"
                    )
                except (TelegramBadRequest, TelegramAPIError) as send_err:
                    logger.warning("HTML send failed in OAuth callback, retrying plain: %s", send_err)
                    try:
                        await self.bot.send_message(
                            telegram_id,
                            _strip_html_tags(message_text),
                        )
                    except Exception as inner:
                        logger.error("Plain-text fallback also failed in OAuth callback: %s", inner)
            return "✅ Ruxsat olindi! Endi havolani yuboring."
        
        except ValueError:
            # Re-raise ValueError so _google_callback can handle state errors
            raise
        except Exception as e:
            logger.error(f"❌ OAuth callback error: {e}", exc_info=True)
            return f"❌ OAuth xatoligi: {str(e)[:100]}"


class OAuthServer:
    def __init__(self, context: AppContext) -> None:
        self.context = context
        self.runner: web.AppRunner | None = None
        self.site: web.TCPSite | None = None

    async def start(self) -> None:
        app = web.Application()
        # Add routes for both Google and GitHub OAuth callbacks
        app.add_routes([
            web.get("/", self._google_callback),  # Google OAuth callback
            web.get("/github/callback", self._github_callback),  # GitHub OAuth callback
            web.get("/health", self._health_check),  # Health check endpoint
        ])
        self.runner = web.AppRunner(app)
        await self.runner.setup()
        # Use config values for host and port
        host = self.context.config.server_host
        port = self.context.config.server_port
        self.site = web.TCPSite(self.runner, host=host, port=port)
        await self.site.start()
        logger.info(f"✅ OAuth callback server running at {self.context.config.github_redirect_uri}")

    async def stop(self) -> None:
        if self.runner:
            await self.runner.cleanup()

    async def _health_check(self, request: web.Request) -> web.Response:
        """Health check endpoint - returns bot version and status"""
        return web.json_response({
            "status": "🟢 OK",
            "version": BOT_VERSION,
            "features": FEATURES,
            "timestamp": datetime.now().isoformat(),
            "ai_qa_enabled": FEATURES.get("ai_qa", False),
            "data_indexing_enabled": FEATURES.get("data_indexing", False),
        })

    async def _google_callback(self, request: web.Request) -> web.Response:
        """Handle Google OAuth callback"""
        state = request.query.get("state")
        code = request.query.get("code")
        error = request.query.get("error")
        if error:
            return web.Response(
                text=f"OAuth xatoligi: {error}. Telegram botga qayting va qayta urinib ko'ring."
            )
        if not state or not code:
            return web.Response(text="Noto'g'ri OAuth so'rovi.")
        try:
            msg = await self.context.handle_oauth_callback(state, code)
            return web.Response(text=msg)
        except ValueError as ve:
            error_str = str(ve)
            logger.warning(f"⚠️ OAuth state error: {error_str}")
            # Provide helpful message for state errors
            if "state not in" in error_str.lower() or "holati topilmadi" in error_str.lower():
                return web.Response(
                    text="❌ OAuth sessiyasi muddati tugagan.\n\n"
                         "Iltimos, Telegram botga qaytib /start buyrug'ini yuboring va qayta urinib ko'ring.\n\n"
                         "Agar muammo davom etsa, bir necha daqiqaga qayta urinib ko'ring."
                )
            else:
                return web.Response(text=f"❌ OAuth xatosi: {error_str[:80]}")
        except Exception as exc:
            logger.exception("OAuth callback xatosi: %s", exc)
            return web.Response(
                text=f"❌ Xato: {str(exc)[:100]}.\n\n"
                     "Iltimos, /start'ga qaytib qayta urinib ko'ring."
            )

    async def _github_callback(self, request: web.Request) -> web.Response:
        """Handle GitHub OAuth callback"""
        code = request.query.get("code")
        state = request.query.get("state")
        error = request.query.get("error")
        
        if error:
            error_desc = request.query.get("error_description", error)
            return web.Response(
                text=f"GitHub OAuth xatoligi: {error_desc}. Telegram botga qayting va qayta urinib ko'ring."
            )
        
        if not code or not state:
            return web.Response(text="Noto'g'ri GitHub OAuth so'rovi.")
        
        try:
            result = await self.context.github_oauth_service.exchange_code_for_token(code, state)
            
            if not result:
                return web.Response(text="❌ GitHub autentifikatsiyasi muvaffaq bo'lmadi. Qayta urinib ko'ring.")
            
            telegram_id = result["telegram_id"]
            github_username = result.get("github_username")
            github_email = result.get("github_email")
            github_token = result.get("access_token")
            
            # Save GitHub credentials to database
            try:
                await asyncio.to_thread(
                    self.context._save_github_credentials_sync,
                    telegram_id,
                    github_username,
                    github_email,
                    github_token
                )
            except Exception as e:
                logger.warning(f"⚠️ Could not save GitHub credentials: {e}")
            
            # Send success message to user
            if self.context.bot:
                message_text = f"✅ <b>GitHub hisobiga muvaffaqiyatli ulandi!</b>\n\n" \
                              f"👤 <b>Foydalanuvchi:</b> @{html_escape(github_username)}\n" \
                              f"📧 <b>Email:</b> {html_escape(github_email or 'Nomalum')}\n\n" \
                              f"Endi siz barcha xususiyatlardan foydalanishingiz mumkin."
                try:
                    await self.context.bot.send_message(
                        telegram_id,
                        message_text,
                        parse_mode="HTML"
                    )
                    # Send main menu
                    await self.context.bot.send_message(
                        telegram_id,
                        "🏠 Asosiy menyu:",
                        reply_markup=build_main_menu()
                    )
                except Exception as e:
                    logger.warning(f"⚠️ Could not send message to user: {e}")
            
            return web.Response(text="✅ GitHub autentifikatsiyasi muvaffaqiyatli! Telegramga qayting.")
        
        except Exception as exc:
            logger.exception("GitHub callback xatosi: %s", exc)
            return web.Response(
                text=f"Xato: {str(exc)[:100]}. Qayta urinib ko'ring."
            )


def register_handlers(dp: Dispatcher, ctx: AppContext) -> None:
    @dp.message(CommandStart())
    async def start_handler(message: Message) -> None:
        """Handle /start command"""
        telegram_id = message.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"📱 /start - User {telegram_id}")
        
        try:
            # Check if user already registered
            logger.info(f"🔍 Checking if user {telegram_id} is registered in Supabase...")
            user = await ctx.supabase_service.get_user_by_telegram(telegram_id)
            logger.info(f"📊 Supabase result: {user}")
            
            if user:
                # User already registered - show main menu
                first_name = user.get("first_name", "")
                last_name = user.get("last_name", "")
                full_name = f"{first_name} {last_name}".strip()
                
                session.full_name = full_name
                session.email = user.get("email")

                # session.step may already be "in_chat" if workspace was restored
                # by SessionStore._restore_workspace(); don't overwrite it.
                if session.step not in ("in_chat", "selecting_folder_sheets"):
                    session.step = "ready"
                
                # Check for active sheet integration (Supabase metadata only)
                active = await ctx.supabase_service.get_active_integration(telegram_id)
                if active:
                    if not session.sheet_id:
                        session.sheet_id = active.get("sheet_id")
                    if not session.sheet_name:
                        session.sheet_name = active.get("sheet_name")
                
                logger.info(f"✅ User {telegram_id} already registered: {full_name}")

                # Build a greeting that reflects the actual state
                if session.step == "in_chat":
                    # Workspace already loaded — tell the user they can go straight to chat
                    if session.all_folder_sheets_data:
                        n = len(session.all_folder_sheets_data)
                        ws_hint = f"📁 {n} ta spreadsheet yuklangan. Savol berishingiz mumkin!"
                    elif session.all_sheets_data:
                        n = len(session.all_sheets_data)
                        ws_hint = f"📊 {session.sheet_name or 'Spreadsheet'} ({n} ta varaq) yuklangan. Savol berishingiz mumkin!"
                    else:
                        ws_hint = "Menyudan kerakli bo'limni tanlang:"
                        session.step = "ready"
                    await message.answer(
                        f"👋 Xush kelibsiz, {full_name}!\n\n{ws_hint}",
                        reply_markup=build_main_menu(),
                    )
                else:
                    await message.answer(
                        f"👋 Xush kelibsiz, {full_name}!\n\n"
                        "Menyudan kerakli bo'limni tanlang:",
                        reply_markup=build_main_menu(),
                    )
                logger.info(f"✅ Response sent to {telegram_id}")
                return
            
            # New user - start registration
            logger.info(f"🆕 New user {telegram_id} - starting registration")
            session.step = "waiting_first_name"
            
            logger.info(f"📤 Sending registration prompt to {telegram_id}...")
            await message.answer(
                "Assalomu alaykum! 👋 OnBrain AI botiga xush kelibsiz.\n\n"
                "✍️ <b>Ro'yxatdan o'tish uchun, iltimos, ismingizni yuboring:</b>",
                parse_mode="HTML",
            )
            logger.info(f"✅ Registration prompt sent to {telegram_id}")
            
        except Exception as exc:
            logger.exception(f"❌ /start handler error for user {telegram_id}: {exc}")
            try:
                await message.answer(
                    "❌ Xatolik yuz berdi. Iltimos, /start buyrug'ini qayta yuboring."
                )
            except Exception as e:
                logger.error(f"❌ Could not send error message: {e}")

    @dp.message(Command("chat"))
    async def chat_command_handler(message: Message) -> None:
        """Handle /chat command to start chatting"""
        telegram_id = message.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"💬 /chat command - User {telegram_id}")
        
        # Check if user is registered
        if session.step in {"waiting_first_name", "waiting_last_name", "waiting_contact", "waiting_email"}:
            await message.answer("❌ Avval ro'yxatdan o'tishni yakunlang. /start buyrug'ini yuboring.")
            return
        
        try:
            session.step = "in_chat"
            await message.answer(
                "💬 <b>Chat Mode</b>\n\n"
                "Assalomu alaykum! Men OnBrain AI yordamchisiman.\n\n"
                "Istalgan savolingizni yuboring va men sizga javob beraman.\n\n"
                "<i>Misollar:</i>\n"
                "• Python nima?\n"
                "• Machine Learning qanday ishlaydi?\n"
                "• Uzbekistonda qaysi university eng yaxshi?\n\n"
                "Suhbatni tugatish uchun /start buyrug'ini yuboring.",
                parse_mode="HTML"
            )
        except Exception as exc:
            logger.exception(f"❌ Chat command error: {exc}")
            await message.answer("❌ Xatolik yuz berdi. Iltimos, qayta urinib ko'ring.")

    @dp.message(Command("sheets"))
    async def sheets_command_handler(message: Message) -> None:
        """Handle /sheets command to connect Google Sheets"""
        telegram_id = message.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"📊 /sheets command - User {telegram_id}")
        
        # Check if user is registered
        if session.step in {"waiting_first_name", "waiting_last_name", "waiting_contact", "waiting_email"}:
            await message.answer("❌ Avval ro'yxatdan o'tishni yakunlang. /start buyrug'ini yuboring.")
            return
        
        try:
            # Trigger sheets button handler
            session.auth_mode = "sheets"
            try:
                await message.answer(
                    "📊 <b>Google Sheets ulash</b>\n\n"
                    "Google Sheets fayli ulash uchun:\n\n"
                    "1️⃣ Google hisobiga kiring\n"
                    "2️⃣ Spreadsheet linkini yuboring:\n"
                    "<code>https://docs.google.com/spreadsheets/d/1ABC123xyz/edit</code>",
                    parse_mode="HTML"
                )
            except (TelegramBadRequest, TelegramAPIError) as exc:
                logger.warning("/sheets command HTML send failed, sending plain: %s", exc)
                await message.answer(
                    "📊 Google Sheets ulash\n\n"
                    "Google Sheets fayli ulash uchun:\n\n"
                    "1. Google hisobiga kiring\n"
                    "2. Spreadsheet linkini yuboring:\n"
                    "https://docs.google.com/spreadsheets/d/1ABC123xyz/edit"
                )
            
            # Try to load & auto-refresh token from SQLite/Supabase (works after restart too)
            if not session.google_credentials_json:
                refreshed = await asyncio.to_thread(load_and_refresh_google_token, telegram_id)
                if refreshed:
                    session.google_credentials_json = refreshed
                    logger.info(f"✅ Force-loaded & refreshed Google token for user {telegram_id}")
            
            # Now check again after reload attempt
            if not session.google_credentials_json:
                session.step = "waiting_auth"
                await message.answer(
                    "🔐 Iltimos, Google hisobiga kiring:",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="🔐 Google'da kirish", callback_data="auth_google")]
                    ])
                )
        except Exception as exc:
            logger.exception(f"❌ Sheets command error: {exc}")
            await message.answer("❌ Xatolik yuz berdi. Iltimos, qayta urinib ko'ring.")

    @dp.message(Command("folder"))
    async def folder_command_handler(message: Message) -> None:
        """Handle /folder command to access Google Drive folder"""
        telegram_id = message.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"📁 /folder command - User {telegram_id}")
        
        # Check if user is registered
        if session.step in {"waiting_first_name", "waiting_last_name", "waiting_contact", "waiting_email"}:
            await message.answer("❌ Avval ro'yxatdan o'tishni yakunlang. /start buyrug'ini yuboring.")
            return
        
        try:
            session.auth_mode = "folder"
            try:
                await message.answer(
                    "📁 <b>Google Drive Papka ulash</b>\n\n"
                    "Google Drive papka uchun:\n\n"
                    "1️⃣ Google hisobiga kiring\n"
                    "2️⃣ Papka linkini yuboring:\n"
                    "<code>https://drive.google.com/drive/folders/1ABC123xyz</code>",
                    parse_mode="HTML"
                )
            except (TelegramBadRequest, TelegramAPIError) as exc:
                logger.warning("/folder command HTML send failed, sending plain: %s", exc)
                await message.answer(
                    "📁 Google Drive Papka ulash\n\n"
                    "Google Drive papka uchun:\n\n"
                    "1. Google hisobiga kiring\n"
                    "2. Papka linkini yuboring:\n"
                    "https://drive.google.com/drive/folders/1ABC123xyz"
                )
            
            # Try to load & auto-refresh token from SQLite/Supabase (works after restart too)
            if not session.google_credentials_json:
                refreshed = await asyncio.to_thread(load_and_refresh_google_token, telegram_id)
                if refreshed:
                    session.google_credentials_json = refreshed
                    logger.info(f"✅ Force-loaded & refreshed Google token for user {telegram_id}")
            
            # Now check again after reload attempt
            if not session.google_credentials_json:
                session.step = "waiting_auth"
                await message.answer(
                    "🔐 Iltimos, Google hisobiga kiring:",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="🔐 Google'da kirish", callback_data="auth_google")]
                    ])
                )
        except Exception as exc:
            logger.exception(f"❌ Folder command error: {exc}")
            await message.answer("❌ Xatolik yuz berdi. Iltimos, qayta urinib ko'ring.")

    @dp.message(Command("excel"))
    async def excel_command_handler(message: Message) -> None:
        """Handle /excel command to upload Excel file"""
        telegram_id = message.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"📁 /excel command - User {telegram_id}")
        
        # Check if user is registered
        if session.step in {"waiting_first_name", "waiting_last_name", "waiting_contact", "waiting_email"}:
            await message.answer("❌ Avval ro'yxatdan o'tishni yakunlang. /start buyrug'ini yuboring.")
            return
        
        try:
            session.step = "waiting_excel_file"
            await message.answer(
                "📁 <b>Excel fayl yuklash</b>\n\n"
                "Excel fayl (.xlsx, .xls) yuboring:\n\n"
                "Botni fayl bilan javob bering.",
                parse_mode="HTML"
            )
        except Exception as exc:
            logger.exception(f"❌ Excel command error: {exc}")
            await message.answer("❌ Xatolik yuz berdi. Iltimos, qayta urinib ko'ring.")

    @dp.message(Command("disconnect"))
    async def disconnect_command_handler(message: Message) -> None:
        """Disconnect the user's Google Drive/Sheets workspace and clear cached data."""
        telegram_id = message.from_user.id
        session = ctx.sessions.get(telegram_id)
        logger.info(f"🔌 /disconnect - User {telegram_id}")

        if session.step in {"waiting_first_name", "waiting_last_name", "waiting_contact", "waiting_email"}:
            await message.answer("❌ Avval ro'yxatdan o'tishni yakunlang.")
            return

        # Clear session data
        session.sheet_id = None
        session.sheet_name = None
        session.sheet_data = []
        session.all_sheets_data = {}
        session.all_folder_sheets_data = {}
        session.folder_spreadsheets = []
        session.selected_spreadsheets = []
        session.folder_id = None
        session.pending_sheets = {}
        session.step = "ready"

        # Remove from SQLite workspace store
        await asyncio.to_thread(_workspace_store.delete_workspace, telegram_id)

        await message.answer(
            "✅ <b>Ulanish uzildi.</b>\n\n"
            "Barcha spreadsheet ma'lumotlari o'chirildi.\n\n"
            "Qayta ulash uchun /sheets yoki /folder buyrug'ini yuboring.",
            parse_mode="HTML",
            reply_markup=build_main_menu(),
        )
        logger.info(f"✅ Workspace disconnected for user {telegram_id}")

    @dp.message(Command("help"))
    async def _help_command_handler(message: Message) -> None:
        """Handle /help command to show help information"""
        telegram_id = message.from_user.id
        logger.info(f"❓ /help command - User {telegram_id}")
        
        try:
            help_text = (
                "<b>📚 OnBrain AI Bot - Yordam</b>\n\n"
                "<b>Mavjud buyruqlar:</b>\n"
                "🏠 <b>/start</b> - Asosiy menyu\n"
                "💬 <b>/chat</b> - Assistant bilan suhbat\n"
                "📊 <b>/sheets</b> - Google Sheets ulash\n"
                "📁 <b>/folder</b> - Google Drive Papka ulash\n"
                "📄 <b>/excel</b> - Excel fayl yuklash\n"
                "🔌 <b>/disconnect</b> - Ulangan spreadsheetni uzish\n"
                "❓ <b>/help</b> - Bu yordam matnini ko'rsatish\n\n"
                "<b>Asosiy funksiyalar:</b>\n"
                "📊 <b>Google Sheets</b> - Google Sheets fayllari bilan ishlash\n"
                "📁 <b>Google Drive</b> - Google Drive papkasidagi fayllari o'qish\n"
                "📄 <b>Excel</b> - Excel fayllari yuklab jo'natish\n"
                "💬 <b>Chat</b> - AI yordamchi bilan suhbat\n\n"
                "<b>Qanday ishlatish:</b>\n"
                "1. /start buyrug'ini yuboring\n"
                "2. Asosiy menyu dan kerakli bo'limni tanlang\n"
                "3. Google Sheets yoki Excel fayl jo'nating\n"
                "4. Savolingizni yozing va javob oling!\n\n"
                "<i>Agar muammo bo'lsa, @aionbrain_bot ga murojaat qiling</i>"
            )
            await message.answer(help_text, parse_mode="HTML")
        except Exception as exc:
            logger.exception(f"❌ Help command error: {exc}")
            await message.answer("❌ Xatolik yuz berdi. Iltimos, qayta urinib ko'ring.")

    @dp.callback_query(F.data == "chat_start")
    async def chat_button_handler(callback_query: CallbackQuery) -> None:
        """Handle Chat button click"""
        telegram_id = callback_query.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"💬 Chat button clicked - User {telegram_id}")
        
        # Check if user is registered
        if session.step in {"waiting_first_name", "waiting_last_name", "waiting_contact", "waiting_email"}:
            await callback_query.answer("❌ Avval ro'yxatdan o'tishni yakunlang!", show_alert=True)
            return
        
        try:
            session.step = "in_chat"
            await callback_query.message.answer(
                "💬 <b>Chat Mode</b>\n\n"
                "Assalomu alaykum! Men OnBrain AI yordamchisiman.\n\n"
                "Istalgan savolingizni yuboring va men sizga javob beraman.\n\n"
                "<i>Misollar:</i>\n"
                "• Python nima?\n"
                "• Machine Learning qanday ishlaydi?\n"
                "• Uzbekistonda qaysi university eng yaxshi?\n\n"
                "💡 <b>Maslahat:</b> Quyidagi tugmani bosing yoki /start buyrug'ini yuboring chat rejimini tark etish uchun.",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="🚪 Chat'ni tark etish", callback_data="exit_chat")]
                ])
            )
            await callback_query.answer("✅ Chat rejimi faollashtirildi")
        except Exception as exc:
            logger.exception(f"❌ Chat button error: {exc}")
            await callback_query.answer("❌ Xatolik yuz berdi!", show_alert=True)

    @dp.callback_query(F.data == "exit_chat")
    async def exit_chat_handler(callback_query: CallbackQuery) -> None:
        """Handle Exit Chat button click"""
        telegram_id = callback_query.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"🚪 Exit chat clicked - User {telegram_id}")
    
        try:
            session.step = "ready"
            await callback_query.answer("✅ Chat rejimi yopildi")
            
            # Show main menu with all options
            await callback_query.message.answer(
                "👋 <b>Asosiy menyu</b>\n\n<i>Nimani qilmoqchisiz?</i>",
                reply_markup=build_main_menu(),
                parse_mode="HTML"
            )
        except Exception as exc:
            logger.exception(f"❌ Exit chat error: {exc}")
            await callback_query.answer("❌ Xatolik yuz berdi!", show_alert=True)

    @dp.callback_query(F.data == "chat_continue")
    async def chat_continue_handler(callback_query: CallbackQuery) -> None:
        """Handle Continue Chat button click - user wants to ask more questions"""
        telegram_id = callback_query.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"💬 Continue chat clicked - User {telegram_id}")
        
        try:
            session.step = "in_chat"
            session.web_search_mode = False  # Reset web search mode
            await callback_query.answer("✅ Chat davom etmoqda...", show_alert=False)
            await callback_query.message.answer(
                "💬 Yana savolingizni yozing yoki /start buyrug'ini bosing asosiy menuyga qaytish uchun."
            )
        except Exception as exc:
            logger.exception(f"❌ Chat continue error: {exc}")
            await callback_query.answer("❌ Xatolik yuz berdi!", show_alert=True)

    @dp.callback_query(F.data == "web_search_mode")
    async def web_search_mode_handler(callback_query: CallbackQuery) -> None:
        """Switch to internet web search mode (Tavily), bypasses spreadsheet data"""
        telegram_id = callback_query.from_user.id
        session = ctx.sessions.get(telegram_id)
        session.step = "in_chat"
        session.web_search_mode = True
        await callback_query.answer("🌐 Internet qidirish yoqildi", show_alert=False)
        await callback_query.message.answer(
            "🌐 <b>Internet qidirish rejimi</b>\n\n"
            "Endi savolingizni yozing — men internetdan qidiraman.\n"
            "Spreadsheet ma'lumotlariga qaytish uchun <b>Chat-ga qaytish</b> tugmasini bosing.",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📊 Spreadsheet rejimiga qaytish", callback_data="chat_continue")],
                [InlineKeyboardButton(text="🚪 Chat-ni tugatish", callback_data="exit_chat")],
            ])
        )

    @dp.callback_query(F.data == "main_menu")
    async def main_menu_handler(callback_query: CallbackQuery) -> None:
        """Handle Main Menu button - return to main menu"""
        telegram_id = callback_query.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"🏠 Main menu clicked - User {telegram_id}")
        
        try:
            session.step = "ready"
            await callback_query.answer("✅ Asosiy menyu")
            
            # Show main menu with all options
            await callback_query.message.answer(
                "👋 <b>Asosiy menyu</b>\n\n<i>Nimani qilmoqchisiz?</i>",
                reply_markup=build_main_menu(),
                parse_mode="HTML"
            )
        except Exception as exc:
            logger.exception(f"❌ Main menu error: {exc}")
            await callback_query.answer("❌ Xatolik yuz berdi!", show_alert=True)

    @dp.callback_query(F.data == "sheets")
    async def sheets_button_handler(callback_query: CallbackQuery) -> None:
        """Handle Google Sheets button click from inline menu"""
        telegram_id = callback_query.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"📊 Sheets button clicked - User {telegram_id}")
        
        try:
            # Set auth mode to "sheets" so OAuth knows what to do
            session.auth_mode = "sheets"
            
            # Call the same logic as text handler
            if session.step in {"waiting_name", "waiting_email"}:
                await callback_query.answer("Avval ro'yxatdan o'tishni yakunlang. /start buyrug'ini yuboring.", show_alert=True)
                return
            
            # Check if user already has Google credentials
            if session.google_credentials_json:
                # User already authenticated, ask for sheet link
                try:
                    await callback_query.message.edit_text(
                        "📊 <b>Google Sheets ulash</b>\n\n"
                        "✅ Siz Google hisobiga ulangansiz!\n\n"
                        "Endi Google Sheets havolasini jo'nating:\n\n"
                        "1️⃣ Google Sheets ochib, shunga daxl qiling\n"
                        "2️⃣ \"Ulashish\" (Share) tugmasini bosing\n"
                        "3️⃣ \"Qoʻl kiritish\" (Anyone with link) tanlang\n"
                        "4️⃣ Havolani nusxa olib, bot ga yuboring\n\n"
                        "📋 <b>Misol:</b>\n"
                        "<code>https://docs.google.com/spreadsheets/d/1Abc123xyz/edit</code>",
                        parse_mode="HTML"
                    )
                except (TelegramBadRequest, TelegramAPIError) as exc:
                    logger.warning("Sheets button edit_text failed, sending plain: %s", exc)
                    await callback_query.message.answer(
                        "📊 Google Sheets ulash\n\n"
                        "✅ Siz Google hisobiga ulangansiz!\n\n"
                        "Endi Google Sheets havolasini jo'nating:\n\n"
                        "Misol:\nhttps://docs.google.com/spreadsheets/d/1Abc123xyz/edit"
                    )
                session.step = "waiting_sheet_link"
                await callback_query.message.answer("📌 Google Sheets havolasini yuboring...")
            else:
                # User not authenticated, require Google login first
                await callback_query.message.edit_text(
                    "🔐 <b>Google Security</b>\n\n"
                    "Xavfsizlik uchun avval Google hisobiga ulanishingiz kerak.\n\n"
                    "Quyidagi havolani oching va ruxsat bering:",
                    parse_mode="HTML"
                )
                
                auth_url = ctx.oauth_service.create_auth_url(telegram_id)
                safe_url = _escape_url_for_html(auth_url)
                try:
                    await callback_query.message.answer(
                        f"🔗 <a href='{safe_url}'>Google bilan kirish</a>\n\n"
                        "Ruxsat berganingizdan keyin, bot Google Sheets ulashishni taklif qiladi.",
                        parse_mode="HTML"
                    )
                except (TelegramBadRequest, TelegramAPIError) as exc:
                    logger.warning("Auth URL HTML send failed, sending plain: %s", exc)
                    await callback_query.message.answer(
                        f"🔗 Google bilan kirish:\n{auth_url}\n\n"
                        "Ruxsat berganingizdan keyin, bot Google Sheets ulashishni taklif qiladi."
                    )
                
                session.step = "waiting_google_auth"
                await callback_query.message.answer(
                    "⏳ Ruxsat berganingizdan keyin, \"📊 Google Sheets ulash\" tugmasini yana bosing..."
                )
            
            await callback_query.answer("✅ Sheets bo'limi faollashtirildi")
        except Exception as exc:
            logger.exception(f"❌ Sheets button error: {exc}")
            await callback_query.answer("❌ Xatolik yuz berdi!", show_alert=True)

    @dp.callback_query(F.data == "excel")
    async def excel_button_handler(callback_query: CallbackQuery) -> None:
        """Handle Excel button click from inline menu"""
        telegram_id = callback_query.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"📁 Excel button clicked - User {telegram_id}")
        
        try:
            if session.step in {"waiting_name", "waiting_email"}:
                await callback_query.answer("Avval ro'yxatdan o'tishni yakunlang. /start buyrug'ini yuboring.", show_alert=True)
                return
            
            session.step = "waiting_excel"
            await callback_query.message.edit_text(
                "Excel fayl yuboring (.xlsx yoki .xls).\n"
                "Fayl qabul qilingach, savollaringizga ma'lumot asosida javob beraman."
            )
            await callback_query.answer("✅ Excel bo'limi faollashtirildi")
        except Exception as exc:
            logger.exception(f"❌ Excel button error: {exc}")
            await callback_query.answer("❌ Xatolik yuz berdi!", show_alert=True)

    @dp.callback_query(F.data == "folder")
    async def folder_button_handler(callback_query: CallbackQuery) -> None:
        """Handle Google Drive Folder button click"""
        telegram_id = callback_query.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"📁 Folder button clicked - User {telegram_id}")
        
        try:
            if session.step in {"waiting_name", "waiting_email"}:
                await callback_query.answer("Avval ro'yxatdan o'tishni yakunlang. /start buyrug'ini yuboring.", show_alert=True)
                return
            
            # Set auth mode to "folder" so OAuth knows what to do
            session.auth_mode = "folder"
            
            # Check if user already has Google credentials
            if session.google_credentials_json:
                # User already authenticated, ask for folder link
                try:
                    await callback_query.message.edit_text(
                        "📁 <b>Google Drive Papka ulash</b>\n\n"
                        "✅ Siz Google hisobiga ulangansiz!\n\n"
                        "Endi Google Drive papka havolasini jo'nating:\n\n"
                        "1️⃣ Google Drive ni oching (drive.google.com)\n"
                        "2️⃣ Spreadsheet lar joylashgan papkani toping\n"
                        "3️⃣ Papka ustiga o'ng click → \"Ulashish\" (Share) tugmasini bosing\n"
                        "4️⃣ \"Qoʻl kiritish\" (Anyone with link) tanlang\n"
                        "5️⃣ Havolani nusxa olib, bot ga yuboring\n\n"
                        "📋 <b>Misol:</b>\n"
                        "<code>https://drive.google.com/drive/folders/1ABC123xyz</code>",
                        parse_mode="HTML"
                    )
                except (TelegramBadRequest, TelegramAPIError) as exc:
                    logger.warning("Folder button edit_text failed, sending plain: %s", exc)
                    await callback_query.message.answer(
                        "📁 Google Drive Papka ulash\n\n"
                        "✅ Siz Google hisobiga ulangansiz!\n\n"
                        "Endi Google Drive papka havolasini jo'nating:\n\n"
                        "Misol:\nhttps://drive.google.com/drive/folders/1ABC123xyz"
                    )
                session.step = "waiting_folder_link"
                await callback_query.message.answer("📌 Google Drive papka havolasini yuboring...")
            else:
                # User not authenticated, require Google login first
                await callback_query.message.edit_text(
                    "🔐 <b>Google Security</b>\n\n"
                    "Xavfsizlik uchun avval Google hisobiga ulanishingiz kerak.\n\n"
                    "Quyidagi havolani oching va ruxsat bering:",
                    parse_mode="HTML"
                )
                
                auth_url = ctx.oauth_service.create_auth_url(telegram_id)
                safe_url = _escape_url_for_html(auth_url)
                try:
                    await callback_query.message.answer(
                        f"🔗 <a href='{safe_url}'>Google bilan kirish</a>\n\n"
                        "Ruxsat berganingizdan keyin, bot Google Drive papka ulashishni taklif qiladi.",
                        parse_mode="HTML"
                    )
                except (TelegramBadRequest, TelegramAPIError) as exc:
                    logger.warning("Auth URL HTML send failed in folder handler: %s", exc)
                    await callback_query.message.answer(
                        f"🔗 Google bilan kirish:\n{auth_url}\n\n"
                        "Ruxsat berganingizdan keyin, bot Google Drive papka ulashishni taklif qiladi."
                    )
                
                session.step = "waiting_google_auth"
                await callback_query.message.answer(
                    "⏳ Ruxsat berganingizdan keyin, \"📁 Google Drive Folder\" tugmasini yana bosing..."
                )
            
            await callback_query.answer("✅ Folder bo'limi faollashtirildi")
        except Exception as exc:
            logger.exception(f"❌ Folder button error: {exc}")
            await callback_query.answer("❌ Xatolik yuz berdi!", show_alert=True)

    @dp.callback_query(F.data == "retry_sheets")
    async def retry_sheets_handler(callback_query: CallbackQuery) -> None:
        """Handle retry for sheets - ask for sheets link again"""
        telegram_id = callback_query.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"🔄 Retry sheets - User {telegram_id}")
        
        try:
            session.step = "waiting_sheet_link"
            try:
                await callback_query.message.answer(
                    "📊 <b>Google Sheets ulash</b>\n\n"
                    "Iltimos, Google Sheets havolasini yuboring:\n\n"
                    "1️⃣ Google Sheets ochib, shunga daxl qiling\n"
                    "2️⃣ \"Ulashish\" (Share) tugmasini bosing\n"
                    "3️⃣ Havolani nusxa olib, bot ga yuboring\n\n"
                    "📋 <b>Misol:</b>\n"
                    "<code>https://docs.google.com/spreadsheets/d/1Abc123xyz/edit</code>",
                    parse_mode="HTML"
                )
            except (TelegramBadRequest, TelegramAPIError) as exc:
                logger.warning("Retry sheets HTML failed, sending plain: %s", exc)
                await callback_query.message.answer(
                    "📊 Google Sheets ulash\n\n"
                    "Iltimos, Google Sheets havolasini yuboring:\n\n"
                    "Misol:\nhttps://docs.google.com/spreadsheets/d/1Abc123xyz/edit"
                )
            await callback_query.answer()
        except Exception as exc:
            logger.exception(f"❌ Retry sheets error: {exc}")
            await callback_query.answer("❌ Xatolik yuz berdi!", show_alert=True)

    @dp.callback_query(F.data == "retry_folder")
    async def retry_folder_handler(callback_query: CallbackQuery) -> None:
        """Handle retry for folder - ask for folder link again"""
        telegram_id = callback_query.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"🔄 Retry folder - User {telegram_id}")
        
        try:
            session.step = "waiting_folder_link"
            try:
                await callback_query.message.answer(
                    "📁 <b>Google Drive Papka ulash</b>\n\n"
                    "Iltimos, Google Drive papka havolasini yuboring:\n\n"
                    "1️⃣ Google Drive da papkani oching\n"
                    "2️⃣ \"Ulashish\" (Share) tugmasini bosing\n"
                    "3️⃣ Havolani nusxa olib, bot ga yuboring\n\n"
                    "📋 <b>Misol:</b>\n"
                    "<code>https://drive.google.com/drive/folders/1Abc123xyz?usp=sharing</code>",
                    parse_mode="HTML"
                )
            except (TelegramBadRequest, TelegramAPIError) as exc:
                logger.warning("Retry folder HTML failed, sending plain: %s", exc)
                await callback_query.message.answer(
                    "📁 Google Drive Papka ulash\n\n"
                    "Iltimos, Google Drive papka havolasini yuboring:\n\n"
                    "Misol:\nhttps://drive.google.com/drive/folders/1Abc123xyz?usp=sharing"
                )
            await callback_query.answer()
        except Exception as exc:
            logger.exception(f"❌ Retry folder error: {exc}")
            await callback_query.answer("❌ Xatolik yuz berdi!", show_alert=True)

    @dp.callback_query(F.data == "github_login")
    async def github_login_handler(callback_query: CallbackQuery) -> None:
        """Handle GitHub login button click"""
        telegram_id = callback_query.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"🔐 GitHub login clicked - User {telegram_id}")
        
        try:
            # Check if already authenticated
            if session.github_username:
                await callback_query.answer(
                    f"✅ Siz allaqachon GitHub ga ulangansiz: @{session.github_username}",
                    show_alert=True
                )
                return
            
            await callback_query.answer("🔐 GitHub bilan autentifikatsiya boshlandi...")
            
            # Create GitHub auth URL
            auth_url = ctx.github_oauth_service.create_auth_url(telegram_id)
            
            await callback_query.message.answer(
                "🔐 <b>GitHub Security</b>\n\n"
                "Kriptografik xavfsizlik uchun GitHub hisobiga ulanishingiz kerak.\n\n"
                "Quyidagi havolani oching va ruxsat bering:",
                parse_mode="HTML"
            )
            
            safe_url = _escape_url_for_html(auth_url)
            try:
                await callback_query.message.answer(
                    f"🔗 <a href='{safe_url}'>GitHub bilan kirish</a>\n\n"
                    "Ruxsat berganingizdan keyin, bot sizni GitHub foydalanuvchisi sifatida taniydi.",
                    parse_mode="HTML"
                )
            except (TelegramBadRequest, TelegramAPIError) as exc:
                logger.warning("GitHub auth URL HTML send failed: %s", exc)
                await callback_query.message.answer(
                    f"🔗 GitHub bilan kirish:\n{auth_url}\n\n"
                    "Ruxsat berganingizdan keyin, bot sizni GitHub foydalanuvchisi sifatida taniydi."
                )
            
            session.step = "waiting_github_auth"
            
        except Exception as exc:
            logger.exception(f"❌ GitHub login error: {exc}")
            await callback_query.answer("❌ Xatolik yuz berdi! Qayta urinib ko'ring.", show_alert=True)

    @dp.message(F.text == MAIN_MENU_SHEETS)
    async def connect_sheets_handler(message: Message) -> None:
        telegram_id = message.from_user.id
        session = ctx.sessions.get(telegram_id)
        if session.step in {"waiting_name", "waiting_email"}:
            await message.answer("Avval ro'yxatdan o'tishni yakunlang. /start buyrug'ini yuboring.")
            return
        
        # Check if user already has Google credentials
        if session.google_credentials_json:
            # User already authenticated, ask for sheet link
            try:
                await message.answer(
                    "📊 <b>Google Sheets ulash</b>\n\n"
                    "✅ Siz Google hisobiga ulangansiz!\n\n"
                    "Endi Google Sheets havolasini jo'nating:\n\n"
                    "1️⃣ Google Sheets ochib, shunga daxl qiling\n"
                    "2️⃣ \"Ulashish\" (Share) tugmasini bosing\n"
                    "3️⃣ \"Qoʻl kiritish\" (Anyone with link) tanlang\n"
                    "4️⃣ Havolani nusxa olib, bot ga yuboring\n\n"
                    "📋 <b>Misol:</b>\n"
                    "<code>https://docs.google.com/spreadsheets/d/1Abc123xyz/edit</code>",
                    parse_mode="HTML"
                )
            except (TelegramBadRequest, TelegramAPIError) as exc:
                logger.warning("Sheets menu HTML send failed, sending plain: %s", exc)
                await message.answer(
                    "📊 Google Sheets ulash\n\n"
                    "✅ Siz Google hisobiga ulangansiz!\n\n"
                    "Endi Google Sheets havolasini jo'nating:\n\n"
                    "Misol:\nhttps://docs.google.com/spreadsheets/d/1Abc123xyz/edit"
                )
            session.step = "waiting_sheet_link"
            await message.answer("📌 Google Sheets havolasini yuboring...")
        else:
            # User not authenticated, require Google login first
            await message.answer(
                "🔐 <b>Google Security</b>\n\n"
                "Xavfsizlik uchun avval Google hisobiga ulanishingiz kerak.\n\n"
                "Quyidagi havolani oching va ruxsat bering:",
                parse_mode="HTML"
            )
            
            auth_url = ctx.oauth_service.create_auth_url(telegram_id)
            safe_url = _escape_url_for_html(auth_url)
            try:
                await message.answer(
                    f"🔗 <a href='{safe_url}'>Google bilan kirish</a>\n\n"
                    "Ruxsat berganingizdan keyin, bot Google Sheets ulashishni taklif qiladi.",
                    parse_mode="HTML"
                )
            except (TelegramBadRequest, TelegramAPIError) as exc:
                logger.warning("Auth URL HTML send failed in sheets menu: %s", exc)
                await message.answer(
                    f"🔗 Google bilan kirish:\n{auth_url}\n\n"
                    "Ruxsat berganingizdan keyin, bot Google Sheets ulashishni taklif qiladi."
                )
            
            session.step = "waiting_google_auth"
            await message.answer(
                "⏳ Ruxsat berganingizdan keyin, \"Google Sheets ulash\" tugmasini yana bosing..."
            )

    @dp.message(F.text == MAIN_MENU_EXCEL)
    async def upload_excel_menu_handler(message: Message) -> None:
        telegram_id = message.from_user.id
        session = ctx.sessions.get(telegram_id)
        if session.step in {"waiting_name", "waiting_email"}:
            await message.answer("Avval ro'yxatdan o'tishni yakunlang. /start buyrug'ini yuboring.")
            return
        session.step = "waiting_excel"
        await message.answer(
            "Excel fayl yuboring (.xlsx yoki .xls).\n"
            "Fayl qabul qilingach, savollaringizga ma'lumot asosida javob beraman."
        )

    @dp.message(F.text)
    async def text_handler(message: Message) -> None:
        """Unified text message handler for registration flow and chat"""
        telegram_id = message.from_user.id
        
        # ========== SECURITY: Rate Limiting ==========
        if not rate_limiter.is_allowed(telegram_id):
            await message.answer(
                "⏸️ <b>Juda ko'p so'rovlar!</b>\n\n"
                "Iltimos, bir daqiqa kutib turing va qayta urinib ko'ring.\n\n"
                "🔒 Bu sizni xavfsizlik xatera yuz beradigan hujumlardan himoya qiladi.",
                parse_mode="HTML"
            )
            return
        
        session = ctx.sessions.get(telegram_id)
        
        # ========== SECURITY: Input Validation ==========
        user_text = message.text.strip() if message.text else ""
        if not user_text or len(user_text) > 5000:
            await message.answer("❌ Xabar bo'sh yoki juda uzun.")
            return
        
        # IMPORTANT: Check session.step and handle accordingly
        
        # ====== WAITING FOR GOOGLE SHEETS LINK ======
        if session.step == "waiting_sheet_link":
            user_input = input_validator.sanitize_string(user_text)
            
            # Security check: ensure user has Google authentication
            if not session.google_credentials_json:
                await message.answer(
                    "🔐 <b>Xavfsizlik tekshiruvi</b>\n\n"
                    "❌ Avval Google hisobiga ulanishingiz kerak.\n\n"
                    "Asosiy menyu ga qayting va \"📊 Google Sheets ulash\" tugmasini bosing.",
                    parse_mode="HTML"
                )
                session.step = "ready"
                return
            
            # ---------- Extract sheet ID from any supported link format ----------
            sheet_id = _extract_sheet_id(user_input)

            if sheet_id:
                try:
                    # ========== SECURITY: Validate Sheet ID ==========
                    if not input_validator.validate_sheet_id(sheet_id):
                        logger.warning(f"🚨 Invalid sheet ID format from user {telegram_id}: {sheet_id}")
                        await message.answer("❌ Google Sheets ID noto'g'ri formatda.")
                        return
                    sheet_name = "User Shared Sheet"

                    logger.info(f"📊 User {telegram_id} provided Google Sheets link: {sheet_id}")
                    await message.answer("⏳ Google Sheets o'qilmoqda... Iltimos kuting...")

                    try:
                        # Try using Google Sheets API directly (works better with shared/public sheets)
                        logger.info(f"📊 Loading credentials for user {telegram_id}")
                        
                        try:
                            # Use the existing credential helper function that handles refresh
                            creds = credentials_from_json(session.google_credentials_json, telegram_id=telegram_id)
                            logger.info(f"✅ Credentials loaded and refreshed if needed")
                        except ValueError as scope_error:
                            logger.info(f"Old scopes detected for user {telegram_id}: {scope_error}")
                            await message.answer(
                                "🔄 <b>Yangi ruxsatlar kerak</b>\n\n"
                                "Google Sheets/Drive ning naqsh o'zgarib ketgan. Iltimos, qayta avtentifikatsiya qiling.\n\n"
                                "Asosiy menyudan \"📊 Google Sheets ulash\" tugmasini bosing.",
                                parse_mode="HTML",
                                reply_markup=build_main_menu()
                            )
                            session.google_credentials_json = None  # Clear old credentials
                            session.step = "ready"
                            return
                        except Exception as cred_error:
                            logger.error(f"❌ Credential loading error: {cred_error}")
                            raise ValueError(f"Credentials loading failed: {cred_error}") from cred_error

                        # ── Run all blocking Google API calls in a thread ──────────
                        def _read_sheets_sync():
                            svc = build('sheets', 'v4', credentials=creds)
                            sp = svc.spreadsheets().get(spreadsheetId=sheet_id).execute()
                            sp_name = sp.get('properties', {}).get('title', 'Sheet')
                            result_data = {}
                            ok = 0
                            for sh in sp.get('sheets', []):
                                title = sh['properties']['title']
                                try:
                                    safe_range = f"'{title}'" if (" " in title or "'" in title) else title
                                    r = svc.spreadsheets().values().get(
                                        spreadsheetId=sheet_id, range=safe_range
                                    ).execute()
                                    vals = r.get('values', [])
                                    result_data[title] = vals
                                    if vals:
                                        ok += 1
                                except Exception as _e:
                                    logger.warning(f"⚠️ Could not read sheet '{title}': {_e}")
                                    result_data[title] = []
                            return sp_name, result_data, ok

                        logger.info(f"📊 Reading spreadsheet {sheet_id} in background thread...")
                        sheet_name, all_sheets_data, successfully_read = await asyncio.to_thread(_read_sheets_sync)
                        
                        # Check if we successfully read at least one sheet
                        if successfully_read > 0:
                            session.sheet_id = sheet_id
                            session.sheet_name = sheet_name
                            session.all_sheets_data = all_sheets_data
                            session.sheet_data = []
                            session.excel_data = []
                            session.sheet_id = sheet_id
                            session.sheet_name = sheet_name
                            session.all_sheets_data = all_sheets_data
                            session.sheet_data = []
                            session.excel_data = []
                            session.step = "in_chat"
                            
                            # Show summary
                            sheet_summary = "✅ <b>Google Sheets muvaffaqiyatli ulandi!</b>\n\n"
                            sheet_summary += "📊 <b>Barcha jadvallar:</b>\n"
                            for sheet_name_iter, rows in all_sheets_data.items():
                                row_count = len(rows)
                                col_count = len(rows[0]) if rows else 0
                                sheet_summary += f"📋 {html_escape(sheet_name_iter)}: {row_count} qator, {col_count} ustun\n"
                            
                            sheet_summary += "\n💬 Endi savolingizni yozing, jadval ma'lumotlari asosida javob beraman."
                            try:
                                await message.answer(sheet_summary, parse_mode="HTML", reply_markup=build_chat_response_keyboard())
                            except Exception:
                                # Fallback: send without HTML if parsing fails
                                plain_summary = sheet_summary.replace("<b>", "").replace("</b>", "")
                                await message.answer(plain_summary, reply_markup=build_chat_response_keyboard())
                            
                            logger.info(f"✅ Successfully loaded Google Sheets: {list(all_sheets_data.keys())}")
                        else:
                            await message.answer(
                                "❌ Google Sheets o'qilmadi.\n\n"
                                "Sabablari:\n"
                                "• Sheet umumiy (public) yoki jo'natilgan emas\n"
                                "• Link notog'ri\n\n"
                                "💡 Boshqasi: Sheet havolasini boʻlishanishdan oldin:\n"
                                "1. Google Sheets ochib shunga daxl qiling\n"
                                "2. \"Ulashish\" (Share) tugmasini bosing\n"
                                "3. \"Qoʻl kiritish\" (Anyone with link) tanlang\n"
                                "4. Havolani nusxa olib, bot ga yuboring",
                                reply_markup=build_retry_keyboard("sheets")
                            )
                            session.step = "ready"
                    
                    except Exception as e:
                        logger.error(f"❌ Error reading sheets: {e}", exc_info=True)
                        error_msg = str(e).lower()
                        
                        # ---- PUBLIC FALLBACK: If 403/permission, try public CSV export ----
                        if "permission" in error_msg or "forbidden" in error_msg or "403" in error_msg:
                            logger.info(f"🔄 Trying public CSV fallback for sheet {sheet_id}")
                            try:
                                import csv
                                import io as _io
                                from aiohttp import ClientSession as _ClientSession

                                pub_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
                                async with _ClientSession() as _http:
                                    async with _http.get(pub_url, timeout=15) as resp:
                                        if resp.status == 200:
                                            csv_bytes = await resp.read()
                                            csv_text = csv_bytes.decode("utf-8", errors="replace")
                                            reader = csv.reader(_io.StringIO(csv_text))
                                            rows = [row for row in reader]

                                            if rows:
                                                all_sheets_data = {"Sheet1": rows}
                                                session.sheet_id = sheet_id
                                                session.sheet_name = sheet_name
                                                session.all_sheets_data = all_sheets_data
                                                session.sheet_data = []
                                                session.excel_data = []
                                                session.step = "in_chat"

                                                row_count = len(rows)
                                                col_count = len(rows[0]) if rows else 0
                                                summary = (
                                                    "✅ Google Sheets muvaffaqiyatli ulandi! (public rejimda)\n\n"
                                                    f"📋 Sheet1: {row_count} qator, {col_count} ustun\n\n"
                                                    "💬 Endi savolingizni yozing, jadval ma'lumotlari asosida javob beraman."
                                                )
                                                await message.answer(summary, reply_markup=build_chat_response_keyboard())
                                                logger.info(f"✅ Public CSV fallback worked for sheet {sheet_id}")
                                                return  # success — skip error message below
                                            else:
                                                logger.warning("Public CSV returned empty for sheet %s", sheet_id)
                                        else:
                                            logger.warning("Public CSV export returned status %s for sheet %s", resp.status, sheet_id)
                            except Exception as pub_err:
                                logger.warning("Public CSV fallback failed: %s", pub_err)

                            # If we get here, both API and public fallback failed
                            error_text = (
                                "❌ Faylga kirish huquqi yo'q.\n\n"
                                "Iltimos, quyidagilardan birini bajaring:\n\n"
                                "1️⃣ Google Sheets-da \"Ulashish\" (Share) tugmasini bosing\n"
                                "2️⃣ \"Havolani bilgan har kim\" (Anyone with the link) tanlang\n"
                                "3️⃣ \"Ko'ruvchi\" (Viewer) ruxsatini bering\n"
                                "4️⃣ Havolani qayta yuboring\n\n"
                                "💡 Yoki Sheet egasidan ruxsat so'rang."
                            )
                        elif "not found" in error_msg:
                            error_text = "❌ Sheet topilmadi.\n\nHavolani to'g'ri qilib, qayta urinib ko'ring."
                        else:
                            safe_err = html_escape(str(e)[:100])
                            error_text = f"❌ Xatolik: {safe_err}\n\nHavolani to'g'ri qilib, qayta urinib ko'ring."
                        
                        try:
                            await message.answer(
                                error_text,
                                reply_markup=build_retry_keyboard("sheets")
                            )
                        except Exception as send_err:
                            logger.warning("Inner error send failed: %s", send_err)
                            await message.answer(
                                "❌ Xatolik yuz berdi. Qayta urinib ko'ring.",
                                reply_markup=build_retry_keyboard("sheets")
                            )
                        session.step = "ready"
                
                except Exception as exc:
                    logger.exception(f"❌ Error processing sheet link: {exc}", exc_info=True)
                    
                    # Better error message
                    error_str = str(exc).lower()
                    logger.error(f"🔍 Error details: {error_str}")
                    
                    _fallback_ok = False  # flag — set to True if public CSV works

                    if "permission" in error_str or "forbidden" in error_str or "401" in error_str or "403" in error_str:
                        # ---- PUBLIC FALLBACK (outer handler) ----
                        logger.info(f"🔄 Trying public CSV fallback (outer) for sheet {sheet_id}")
                        try:
                            import csv
                            import io as _io
                            from aiohttp import ClientSession as _ClientSession

                            pub_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
                            async with _ClientSession() as _http:
                                async with _http.get(pub_url, timeout=15) as resp:
                                    if resp.status == 200:
                                        csv_bytes = await resp.read()
                                        csv_text = csv_bytes.decode("utf-8", errors="replace")
                                        reader = csv.reader(_io.StringIO(csv_text))
                                        rows = [row for row in reader]

                                        if rows:
                                            all_sheets_data = {"Sheet1": rows}
                                            session.sheet_id = sheet_id
                                            session.sheet_name = sheet_name
                                            session.all_sheets_data = all_sheets_data
                                            session.sheet_data = []
                                            session.excel_data = []
                                            session.step = "in_chat"

                                            row_count = len(rows)
                                            col_count = len(rows[0]) if rows else 0
                                            summary = (
                                                "✅ Google Sheets muvaffaqiyatli ulandi! (public rejimda)\n\n"
                                                f"📋 Sheet1: {row_count} qator, {col_count} ustun\n\n"
                                                "💬 Endi savolingizni yozing, jadval ma'lumotlari asosida javob beraman."
                                            )
                                            await message.answer(summary, reply_markup=build_chat_response_keyboard())
                                            logger.info(f"✅ Public CSV fallback (outer) worked for sheet {sheet_id}")
                                            _fallback_ok = True
                        except Exception as pub_err:
                            logger.warning("Public CSV fallback (outer) failed: %s", pub_err)

                        if not _fallback_ok:
                            error_msg = (
                                "❌ Google Sheet ga kirish huquqi yo'q.\n\n"
                                "Iltimos, quyidagilardan birini bajaring:\n\n"
                                "1️⃣ Google Sheets-da \"Ulashish\" (Share) tugmasini bosing\n"
                                "2️⃣ \"Havolani bilgan har kim\" (Anyone with the link) tanlang\n"
                                "3️⃣ \"Ko'ruvchi\" (Viewer) ruxsatini bering\n"
                                "4️⃣ Havolani qayta yuboring"
                            )
                    elif "not found" in error_str or "404" in error_str:
                        error_msg = (
                            "❌ <b>Sheet topilmadi</b>\n\n"
                            "Havolani tekshiring, to'g'ri emasdir.\n\n"
                            "<b>To'g'ri format:</b>\n"
                            "<code>https://docs.google.com/spreadsheets/d/ABC123/edit</code>"
                        )
                    elif "invalid_grant" in error_str or "credential" in error_str or "expired" in error_str:
                        error_msg = (
                            "❌ <b>Autentifikatsiya xatosi</b>\n\n"
                            "Google hisobiga qayta ulanishingiz kerak.\n\n"
                            "/start buyrug'ini yuboring va \"📊 Google Sheets\" tugmasini bosing."
                        )
                    else:
                        # Log the full error for debugging
                        logger.error(f"📋 Full error trace: {str(exc)}")
                        # Escape HTML special characters in error message (& must be first!)
                        error_details = html_escape(str(exc)[:80])
                        error_msg = (
                            "❌ <b>Xatolik yuz berdi</b>\n\n"
                            f"Xatolik: {error_details}\n\n"
                            "Iltimos, qayta urinib ko'ring yoki bot egasiga xabar bering."
                        )
                    
                    if not _fallback_ok:
                        try:
                            await message.answer(
                                error_msg,
                                reply_markup=build_retry_keyboard(),
                                parse_mode="HTML"
                            )
                        except (TelegramBadRequest, TelegramAPIError) as send_err:
                            logger.warning("Error-msg HTML send failed, sending plain: %s", send_err)
                            await message.answer(
                                _strip_html_tags(error_msg),
                                reply_markup=build_retry_keyboard()
                            )
                    session.step = "ready"
            elif _looks_like_sheets_url(user_input):
                # It looks like a sheets URL but we couldn't extract an ID
                await message.answer(
                    "❌ Linkdan sheet ID ni ajratib ololmadim.\n\n"
                    "Iltimos, link to'g'ri ekanligini tekshiring.\n"
                    "Quyidagi formatlardan birini yuboring:\n\n"
                    "• https://docs.google.com/spreadsheets/d/SHEET_ID/edit\n"
                    "• https://docs.google.com/spreadsheets/d/SHEET_ID/edit?usp=sharing\n"
                    "• Yoki faqat Sheet ID ni yuboring",
                    reply_markup=build_retry_keyboard()
                )
            else:
                await message.answer(
                    "❌ Google Sheets linki topilmadi.\n\n"
                    "📋 Quyidagi formatlardan birini yuboring:\n\n"
                    "• Google Sheets havolasi\n"
                    "• Yoki faqat Sheet ID (masalan: 1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgVE2upms)\n\n"
                    "💡 Sheet ID — link ichida /d/ dan keyingi uzun belgilar.",
                    reply_markup=build_retry_keyboard()
                )
        
        # ====== WAITING FOR GOOGLE DRIVE FOLDER LINK ======
        elif session.step == "waiting_folder_link":
            user_input = input_validator.sanitize_string(user_text)
            
            # Security check: ensure user has Google authentication
            if not session.google_credentials_json:
                await message.answer(
                    "🔐 <b>Xavfsizlik tekshiruvi</b>\n\n"
                    "❌ Avval Google hisobiga ulanishingiz kerak.\n\n"
                    "Asosiy menyu ga qayting va \"📁 Google Drive Folder\" tugmasini bosing.",
                    parse_mode="HTML"
                )
                session.step = "ready"
                return
            
            # Check if it's a Google Drive folder link (accept both /folders/ and /drive/folders/)
            if "drive.google.com" in user_input and ("folders" in user_input or "drive/" in user_input):
                try:
                    logger.info(f"📁 User {telegram_id} provided Google Drive folder link")
                    await message.answer("⏳ Google Drive papka o'qilmoqda... Iltimos kuting. Bu biroz vaqt olishi mumkin...")
                    
                    # Get credentials
                    try:
                        creds = credentials_from_json(session.google_credentials_json, telegram_id=telegram_id)
                    except ValueError as e:
                        logger.info(f"Old scopes detected for user {telegram_id}: {e}")
                        await message.answer(
                            "🔄 <b>Yangi ruxsatlar kerak</b>\n\n"
                            "Google Sheets/Drive ning naqsh o'zgarib ketgan. Iltimos, qayta avtentifikatsiya qiling.\n\n"
                            "Asosiy menyudan \"📊 Google Sheets ulash\" tugmasini bosing.",
                            parse_mode="HTML",
                            reply_markup=build_main_menu()
                        )
                        session.google_credentials_json = None  # Clear old credentials
                        session.step = "ready"
                        return
                    
                    # Import the Google Drive service
                    from google_drive_service import get_all_spreadsheets_from_folder
                    
                    # Get all spreadsheets from folder
                    spreadsheets, error = await get_all_spreadsheets_from_folder(creds, user_input)
                    
                    if error:
                        # Add helpful tips if it's a permission error
                        if "permission" in error.lower() or "forbidden" in error.lower():
                            error += "\n\n💡 <b>Yechim:</b>\n1. Google hisobiga qayta kiring (📊 Google Sheets)\n2. Papka umumiy (public) ekanligini tekshiring\n3. Havolani qayta yuboring"
                        
                        await message.answer(
                            error,
                            parse_mode="HTML",
                            reply_markup=build_retry_keyboard("folder")
                        )
                        session.step = "ready"
                        return
                    
                    if not spreadsheets:
                        await message.answer(
                            "❌ Papkada spreadsheet topilmadi.",
                            reply_markup=build_retry_keyboard("folder")
                        )
                        session.step = "ready"
                        return
                    
                    # ===== NEW: Start indexing process =====
                    logger.info(f"🚀 Starting indexing for user {telegram_id}...")
                    await message.answer("⏳ Ma'lumotlarni indexlashni boshlamoqda... Bu biroz vaqt olishi mumkin (2-5 minut).")
                    
                    # Extract folder ID from URL
                    folder_id = None
                    patterns = [
                        r'drive\.google\.com/drive/(?:u/\d+/)?folders/([a-zA-Z0-9-_]+)',
                        r'drive\.google\.com/(?:drive)?/(?:u/\d+/)?folders/([a-zA-Z0-9-_]+)',
                    ]
                    for pattern in patterns:
                        match = re.search(pattern, user_input)
                        if match:
                            folder_id = match.group(1)
                            break
                    
                    if not folder_id:
                        await message.answer(
                            "❌ Papka ID-ni ajratib ola olmadim. Iltimos, to'g'ri link yuboring.",
                            reply_markup=build_retry_keyboard("folder")
                        )
                        session.step = "ready"
                        return
                    
                    # Store the folder spreadsheets
                    session.folder_id = folder_id          # persist for workspace save
                    session.folder_spreadsheets = spreadsheets
                    session.selected_spreadsheets = []
                    
                    # Show list of spreadsheets
                    folder_summary = f"✅ <b>Google Drive papkasi muvaffaqiyatli ulandi!</b>\n\n"
                    folder_summary += f"📊 <b>Topilgan {len(spreadsheets)} ta spreadsheet:</b>\n\n"
                    
                    # Create buttons for each spreadsheet
                    keyboard_buttons = []
                    for idx, sheet in enumerate(spreadsheets[:20]):  # Limit to 20 for UI
                        sheet_name = sheet['name']
                        # Truncate long names
                        if len(sheet_name) > 30:
                            sheet_name = sheet_name[:27] + "..."
                        folder_summary += f"{idx+1}. 📋 {sheet['name']}\n"
                        keyboard_buttons.append([
                            InlineKeyboardButton(
                                text=f"✓ {sheet_name}",
                                callback_data=f"select_sheet:{idx}"
                            )
                        ])
                    
                    if len(spreadsheets) > 20:
                        folder_summary += f"\n... va {len(spreadsheets) - 20} ta boshqa spreadsheet"
                    
                    folder_summary += "\n\n💡 O'qimoqchi bo'lgan spreadsheet larni tanlang (bir nechta tanlashingiz mumkin):"
                    
                    # Add "Ready" button
                    keyboard_buttons.append([
                        InlineKeyboardButton(text="✅ Tayyor!", callback_data="load_folder_sheets")
                    ])
                    keyboard_buttons.append([
                        InlineKeyboardButton(text="🔄 Qayta yuborish", callback_data="folder")
                    ])
                    
                    keyboard = InlineKeyboardMarkup(inline_keyboard=keyboard_buttons)
                    
                    await message.answer(
                        folder_summary,
                        parse_mode="HTML",
                        reply_markup=keyboard
                    )
                    
                    session.step = "selecting_folder_sheets"
                    logger.info(f"✅ Showed {len(spreadsheets)} spreadsheets to user {telegram_id}")
                    
                except Exception as e:
                    logger.error(f"❌ Error reading folder: {e}", exc_info=True)
                    
                    # Check if it's a permission/authentication error
                    error_msg = str(e).lower()
                    if "permission" in error_msg or "forbidden" in error_msg or "unauthenticated" in error_msg:
                        await message.answer(
                            "🔐 <b>Ruxsatlar muammosi</b>\n\n"
                            "Google hisobiga qayta kiring va yangi ruxsatlarni bering.\n\n"
                            "Asosiy menyudan \"📊 Google Sheets ulash\" tugmasini bosing.",
                            parse_mode="HTML",
                            reply_markup=build_main_menu()
                        )
                    else:
                        await message.answer(
                            f"❌ Google Drive papka o'qilmadi.\n\n"
                            f"<b>Sabablari:</b>\n"
                            f"• Papka umumiy (public) emas\n"
                            f"• Link notog'ri\n"
                            f"• Papkaga kirish huquqi yo'q\n\n"
                            f"✅ <b>Yechim:</b> Havolani to'g'ri qilib, qayta urinib ko'ring!",
                            parse_mode="HTML",
                            reply_markup=build_retry_keyboard("folder")
                        )
                    session.step = "ready"
            else:
                await message.answer(
                    "❌ Bu Google Drive folder linki emas.\n\n"
                    "📁 To'g'ri formatdagi link yuboring:\n"
                    "https://drive.google.com/drive/folders/1ABC123xyz",
                    reply_markup=build_retry_keyboard("folder")
                )
        
        # ====== IN CHAT MODE ======
        elif session.step == "in_chat":
            try:
                user_message = message.text.strip()
                
                if not user_message:
                    await message.answer("❌ Xabar bo'sh bo'lmasligi kerak.")
                    return
                
                # Show typing indicator
                if ctx.bot:
                    await ctx.bot.send_chat_action(message.chat.id, "typing")
                
                logger.info(f"💬 Chat message from {telegram_id}: {user_message[:50]}")
                logger.info(f"📊 Session data check: sheet_id={session.sheet_id}, "
                           f"all_sheets_data={len(session.all_sheets_data) if session.all_sheets_data else 0} sheets, "
                           f"all_folder_sheets_data={len(session.all_folder_sheets_data) if session.all_folder_sheets_data else 0} spreadsheets, "
                           f"excel_data={len(session.excel_data) if session.excel_data else 0} rows")
                
                # ===== NEW: Check if indexing_service is available =====
                if session.indexing_service:
                    logger.info(f"🚀 Using DataIndexingService for user {telegram_id}")
                    
                    # Query from indexed data
                    success, answer = await session.indexing_service.query_index(user_message)
                    
                    if success:
                        response_text = f"🤖 <b>AI Javob (Indexed Data-dan)</b>\n\n{html_escape(answer)}"
                        
                        # Split response if too long
                        try:
                            if len(response_text) > 4000:
                                parts = [response_text[i:i+4000] for i in range(0, len(response_text), 4000)]
                                for i, part in enumerate(parts):
                                    if i == len(parts) - 1:
                                        await message.answer(part, parse_mode="HTML", reply_markup=build_chat_response_keyboard())
                                    else:
                                        await message.answer(part, parse_mode="HTML")
                            else:
                                await message.answer(response_text, parse_mode="HTML", reply_markup=build_chat_response_keyboard())
                        except Exception:
                            # Fallback: send without HTML if parsing fails
                            plain_text = response_text.replace("<b>", "").replace("</b>", "")
                            await message.answer(plain_text, reply_markup=build_chat_response_keyboard())
                        
                        return
                    else:
                        # Fallback to regular response if indexing fails
                        logger.warning(f"⚠️ Indexing query failed: {answer}")
                        await message.answer(f"⚠️ Index-dan javob olishda xato: {html_escape(str(answer))}")
                        return
                # ===== END: DataIndexingService =====
                
                # Check if we have local spreadsheet data (from folder or single sheet)
                local_context = None
                
                # If user explicitly chose web search mode — skip spreadsheet entirely
                if session.web_search_mode:
                    logger.info(f"🌐 Web search mode active for user {telegram_id} — skipping spreadsheet")
                    local_context = None
                else:
                    # Priority 1: Folder sheets data
                    if session.all_folder_sheets_data:
                        local_context = session.all_folder_sheets_data
                        logger.info(f"📁 Using folder sheets data with {len(local_context)} spreadsheets")
                    # Priority 2: Single sheet data
                    elif session.all_sheets_data:
                        local_context = {session.sheet_id or "sheet": session.all_sheets_data}
                        logger.info(f"📊 Using single sheet data: {list(session.all_sheets_data.keys())}")
                    # Priority 3: Excel file data
                    elif session.excel_data:
                        local_context = {"excel": {"Sheet1": session.excel_data}}
                        logger.info(f"📄 Using Excel file data: {len(session.excel_data)} rows")
                    else:
                        logger.info(f"⚠️ No spreadsheet data in session. sheet_id={session.sheet_id}, all_sheets_data={bool(session.all_sheets_data)}, all_folder_sheets_data={bool(session.all_folder_sheets_data)}, excel_data={bool(session.excel_data)}")
                
                # If we have local data, use it for context
                if local_context:
                    try:
                        # Build context from local spreadsheets — read EVERYTHING
                        context_text = ""
                        
                        for sheet_id, sheets in local_context.items():
                            # Get sheet name from folder_spreadsheets if available
                            sheet_name = next(
                                (s['name'] for s in session.folder_spreadsheets if s['id'] == sheet_id),
                                sheet_id
                            ) if session.folder_spreadsheets else (session.sheet_name or "Spreadsheet")
                            
                            context_text += f"=== Spreadsheet: {sheet_name} ===\n"
                            
                            for sheet_title, rows in sheets.items():
                                context_text += f"\n--- Sheet: {sheet_title} ---\n"
                                if not rows:
                                    context_text += "(bo'sh)\n"
                                    continue
                                
                                # Read ALL rows up to limit, with FULL cell values
                                for i, row in enumerate(rows[:MAX_ROWS_FOR_CONTEXT]):
                                    # Filter out completely empty cells at the end
                                    while row and str(row[-1]).strip() == "":
                                        row = row[:-1]
                                    if not row:
                                        continue  # skip completely empty rows
                                    
                                    # Full cell values — no per-cell truncation
                                    cells = [str(x).strip() for x in row[:MAX_COLS_FOR_CONTEXT]]
                                    context_text += f"Row {i+1}: {' | '.join(cells)}\n"
                                
                                total_rows = len(rows)
                                if total_rows > MAX_ROWS_FOR_CONTEXT:
                                    context_text += f"... va yana {total_rows - MAX_ROWS_FOR_CONTEXT} ta qator bor\n"
                                context_text += "\n"
                        
                        # Limit context size for API calls
                        context_text = context_text[:MAX_CHARS_CONTEXT]
                        
                        logger.info(f"📝 Built local context: {len(context_text)} chars")
                        # DEBUG: Log first 500 chars of context to verify data is actually there
                        logger.info(f"📝 Context preview (first 500 chars):\n{context_text[:500]}")
                        logger.info(f"📝 Context preview (last 300 chars):\n{context_text[-300:]}")
                        
                        # ── Use Grok AI (xAI) to answer based on spreadsheet data ──
                        
                        grok_api_key = os.getenv("GROK_API_KEY", "")
                        
                        if grok_api_key:
                            try:
                                logger.info(f"🤖 Sending to Grok AI for spreadsheet Q&A")
                                
                                system_prompt = (
                                    "Sen spreadsheet ma'lumotlarini tahlil qiluvchi AI assistantsan. "
                                    "Senga spreadsheet ma'lumotlari beriladi. Foydalanuvchi savol beradi. "
                                    "QOIDALAR:\n"
                                    "1. FAQAT berilgan spreadsheet ma'lumotlari asosida javob ber.\n"
                                    "2. Javobni ISHONCHLI va ANIQ ber. 'Ehtimol', 'balki', 'bo'lishi mumkin' so'zlarini ISHLATMA.\n"
                                    "3. Har doim spreadsheet nomini va sheet nomini aytib o'tib javob ber. Masalan: 'Spreadsheet: [nom], Sheet: [nom] ma'lumotlariga ko'ra, ...'\n"
                                    "4. Raqamlarni to'g'ri formatlash: 2500000 -> 2,500,000\n"
                                    "5. Agar ma'lumot spreadsheetda TOPILMASA, aniq ayt: 'Bu ma'lumot spreadsheetda mavjud emas.'\n"
                                    "6. Internetdan yoki boshqa manbalardan hech qanday ma'lumot QO'SHMA.\n"
                                    "7. Javobni o'zbek tilida ber.\n"
                                    "8. Qisqa, aniq va to'g'ridan-to'g'ri javob ber.\n"
                                    "9. ISMLARNI AQLLI QIDIRISH: Agar foydalanuvchi 'Yodgorbek' deb so'rasa, lekin spreadsheetda 'Yodgor' bo'lsa — "
                                    "bu BIR XILL ODAM. O'zbek ismlarida -bek, -boy, -jon, -ali, -xon qo'shimchalari tushirilishi yoki qo'shilishi mumkin. "
                                    "Masalan: Yodgorbek=Yodgor, Jasurbek=Jasur, Sardorbek=Sardor, Nilufar=Nilu, Mahkam=Mahkamboy. "
                                    "Shuningdek, kichik/katta harf farqi bo'lmasin, transliteratsiya (lotin/kirill) ham hisobga olinsin. "
                                    "Har doim ENG YAQIN moslikni topishga harakat qil.\n"
                                    "10. JADVAL TUZILISHI: Spreadsheetda ma'lumotlar turli joylarda bo'lishi mumkin — gorizontal, vertikal, jadval ichida jadval. "
                                    "Barcha qatorlar va ustunlarni tekshir. Ma'lumot birinchi qatorda ham, oxirgi qatorda ham bo'lishi mumkin."
                                )
                                
                                user_prompt = (
                                    f"Quyidagi spreadsheet ma'lumotlari berilgan:\n\n{context_text}\n\n"
                                    f"Savol: {user_message}\n\n"
                                    f"Yuqoridagi spreadsheet ma'lumotlariga asoslanib aniq javob ber. "
                                    f"Ismlar to'liq mos kelmasa ham, eng yaqin moslikni top."
                                )
                                
                                logger.info(f"📤 Grok request: system_prompt={len(system_prompt)} chars, user_prompt={len(user_prompt)} chars (context={len(context_text)} chars)")
                                
                                # Try grok-3-mini-fast first, fallback to other models
                                grok_models = ["grok-3-mini-fast", "grok-3-mini", "grok-2-latest"]
                                ai_answer = None
                                last_error = ""
                                
                                import aiohttp as _aiohttp
                                
                                for model_name in grok_models:
                                    try:
                                        logger.info(f"🤖 Trying Grok model: {model_name}")
                                        
                                        grok_payload = {
                                            "model": model_name,
                                            "messages": [
                                                {"role": "system", "content": system_prompt},
                                                {"role": "user", "content": user_prompt},
                                            ],
                                            "temperature": 0.3,
                                            "max_tokens": 2000,
                                        }
                                        grok_headers = {
                                            "Authorization": f"Bearer {grok_api_key}",
                                            "Content-Type": "application/json",
                                        }
                                        
                                        async with _aiohttp.ClientSession() as _grok_http:
                                            async with _grok_http.post(
                                                "https://api.x.ai/v1/chat/completions",
                                                headers=grok_headers,
                                                json=grok_payload,
                                                timeout=_aiohttp.ClientTimeout(total=90),
                                            ) as grok_resp:
                                                if grok_resp.status == 200:
                                                    grok_data = await grok_resp.json()
                                                    ai_answer = grok_data["choices"][0]["message"]["content"]
                                                    logger.info(f"✅ Grok AI ({model_name}) answer: {ai_answer[:100]}...")
                                                    break
                                                else:
                                                    resp_text = await grok_resp.text()
                                                    last_error = f"{model_name}: {grok_resp.status} - {resp_text[:200]}"
                                                    logger.warning(f"⚠️ Grok model {model_name} failed: {last_error}")
                                    except Exception as model_err:
                                        last_error = f"{model_name}: {str(model_err)[:200]}"
                                        logger.warning(f"⚠️ Grok model {model_name} error: {model_err}")
                                
                                if ai_answer:
                                    response_text = f"🤖 AI Javob\n\n{ai_answer}"
                                else:
                                    logger.error(f"❌ All Grok models failed. Last error: {last_error}")
                                    response_text = (
                                        f"🤖 AI Javob\n\n"
                                        f"⚠️ AI xizmatida vaqtinchalik xatolik ({last_error[:100]})\n"
                                        f"Ma'lumotlar:\n{context_text[:2000]}"
                                    )
                            except Exception as grok_err:
                                logger.error(f"❌ Grok AI error: {grok_err}")
                                # Fallback: show raw data summary
                                response_text = (
                                    f"🤖 AI Javob\n\n"
                                    f"⚠️ AI xizmatida vaqtinchalik xatolik.\n"
                                    f"Ma'lumotlar:\n{context_text[:2000]}"
                                )
                        else:
                            # No Grok key — show raw data as before
                            logger.warning("⚠️ GROK_API_KEY not set, showing raw spreadsheet data")
                            response_text = (
                                f"🤖 AI Javob\n\n"
                                f"⚠️ AI kaliti sozlanmagan. Ma'lumotlar:\n{context_text[:2000]}"
                            )
                        
                        # Send response to user
                        try:
                            if len(response_text) > 4000:
                                parts = [response_text[i:i+4000] for i in range(0, len(response_text), 4000)]
                                for i, part in enumerate(parts):
                                    if i == len(parts) - 1:
                                        await message.answer(part, reply_markup=build_chat_response_keyboard())
                                    else:
                                        await message.answer(part)
                            else:
                                await message.answer(response_text, reply_markup=build_chat_response_keyboard())
                        
                        except Exception as ai_error:
                            logger.error(f"❌ Response error: {ai_error}")
                            await message.answer("❌ Javob berishda xatolik. Qayta urinib ko'ring.")
                        
                        return
                    
                    except Exception as context_error:
                        logger.error(f"❌ Context building error: {context_error}")
                        # Fall through to web search
                
                # Fall back to web search if no local data
                logger.info(f"💻 Using web search (no local spreadsheet data)")
                
                # Get response using Tavily API (web search + AI synthesis)
                try:
                    import requests
                    import html
                    
                    tavily_api_key = os.getenv("TAVILY_API_KEY")
                    
                    if not tavily_api_key:
                        await message.answer("❌ Tavily API key topilmadi!")
                        return
                    
                    logger.info(f"🔍 Using Tavily API for user {telegram_id}")
                    
                    # Use Tavily API to search and get AI-synthesized answer
                    tavily_response = requests.post(
                        "https://api.tavily.com/search",
                        json={
                            "api_key": tavily_api_key,
                            "query": user_message,
                            "include_answer": True,
                            "max_results": 5,
                            "include_images": False,
                        },
                        timeout=10
                    )
                    
                    logger.info(f"📊 Tavily response status: {tavily_response.status_code}")
                    
                    if tavily_response.status_code == 200:
                        search_results = tavily_response.json()
                        
                        # Get Tavily's AI-synthesized answer
                        ai_answer = search_results.get("answer")
                        
                        if ai_answer:
                            logger.info(f"✅ Tavily answer received for: {user_message[:30]}")
                            
                            # Translate answer to Uzbek using Google Translate
                            try:
                                logger.info(f"🌐 Translating response to Uzbek...")
                                from google.cloud import translate_v2
                                
                                translate_client = translate_v2.Client()
                                result = translate_client.translate_text(
                                    ai_answer,
                                    source_language_code="en",
                                    target_language_code="uz"
                                )
                                uzbek_answer = result["translatedText"]
                                logger.info(f"✅ Translation successful: {uzbek_answer[:50]}...")
                            except Exception as trans_err:
                                logger.warning(f"⚠️  Translation failed, using original: {trans_err}")
                                # If translation fails, try using simple HTTP-based translator
                                try:
                                    from_text = requests.get(
                                        "https://api.mymemory.translated.net/get",
                                        params={
                                            "q": ai_answer[:500],
                                            "langpair": "en|uz"
                                        },
                                        timeout=5
                                    ).json()
                                    uzbek_answer = from_text.get("responseData", {}).get("translatedText", ai_answer)
                                    logger.info(f"✅ Fallback translation successful")
                                except:
                                    uzbek_answer = ai_answer
                                    logger.warning(f"⚠️  All translations failed, using English")
                            
                            # Decode HTML entities (like &#39; to ')
                            uzbek_answer = html.unescape(uzbek_answer)
                            
                            # Format response with sources
                            response_text = f"🤖 AI Javob\n\n{uzbek_answer}"
                            
                            # Add sources if available
                            if search_results.get("results"):
                                response_text += "\n\n📚 Manbalar:\n"
                                for i, result in enumerate(search_results["results"][:3], 1):
                                    if result.get("title"):
                                        title = html.unescape(result['title'])
                                        response_text += f"{i}. {title}\n"
                            
                            logger.info(f"✅ Response sent to {telegram_id}: {uzbek_answer[:50]}...")
                            
                            exit_keyboard = InlineKeyboardMarkup(
                                inline_keyboard=[
                                    [InlineKeyboardButton(text="🚪 Chat-ni tugatish", callback_data="exit_chat")]
                                ]
                            )
                            
                            # Split response if too long (Telegram limit is 4096)
                            # Send as plain text to avoid Markdown parse errors from
                            # translated text / web content containing special chars
                            try:
                                if len(response_text) > 4000:
                                    parts = [response_text[i:i+4000] for i in range(0, len(response_text), 4000)]
                                    for i, part in enumerate(parts):
                                        if i == len(parts) - 1:
                                            await message.answer(part, reply_markup=exit_keyboard)
                                        else:
                                            await message.answer(part)
                                else:
                                    await message.answer(response_text, reply_markup=exit_keyboard)
                            except Exception as send_err:
                                logger.warning("AI response send failed: %s", send_err)
                                await message.answer("❌ Javob yuborishda xatolik. Qayta urinib ko'ring.")
                        else:
                            await message.answer("❌ Tavily javob bera olmadi. Keyinroq urinib ko'ring!")
                            logger.warning(f"⚠️  Tavily returned no answer for: {user_message[:30]}")
                    else:
                        error_msg = tavily_response.text
                        logger.error(f"❌ Tavily error {tavily_response.status_code}: {error_msg[:100]}")
                        await message.answer(f"❌ Javob olishda xatolik: {tavily_response.status_code}")
                    
                except Exception as e:
                    error_msg = str(e)
                    logger.error(f"❌ Tavily error: {error_msg}")
                    await message.answer("❌ AI javob berishda xatolik. Qayta urinib ko'ring.")
                
                return
            except Exception as exc:
                logger.exception(f"❌ Chat handler error: {exc}")
                await message.answer("❌ Xatolik yuz berdi. Iltimos, qayta urinib ko'ring.")
                return
        
        # ====== WAITING FOR FIRST NAME ======
        if session.step == "waiting_first_name":
            try:
                first_name = input_validator.sanitize_string(message.text.strip())
                
                # ========== SECURITY: Validate Name ==========
                if not input_validator.validate_name(first_name):
                    await message.answer("❌ Ism faqat harflarni o'z ichiga olishi kerak. Iltimos, qayta urinib ko'ring.")
                    return
                
                if len(first_name) < 2:
                    await message.answer("❌ Iltimos, 2 ta harfdan ko'proq ismingiz bilan qayta urinib ko'ring.")
                    return
                
                session.full_name = first_name
                session.step = "waiting_last_name"
                
                logger.info(f"✅ User {telegram_id} entered first name")
                
                await message.answer(
                    "👤 <b>Endi familiyangizni yuboring:</b>",
                    parse_mode="HTML"
                )
                
            except Exception as exc:
                logger.exception(f"❌ Ism kiritishda xato: {exc}")
                await message.answer("❌ Xatolik yuz berdi. Iltimos, qayta urinib ko'ring.")
            return  # IMPORTANT: Return to prevent further processing
        
        # ====== WAITING FOR LAST NAME ======
        if session.step == "waiting_last_name":
            try:
                last_name = input_validator.sanitize_string(message.text.strip())
                
                # ========== SECURITY: Validate Name ==========
                if not input_validator.validate_name(last_name):
                    await message.answer("❌ Familiya faqat harflarni o'z ichiga olishi kerak. Iltimos, qayta urinib ko'ring.")
                    return
                
                if len(last_name) < 2:
                    await message.answer("❌ Iltimos, 2 ta harfdan ko'proq familiyangiz bilan qayta urinib ko'ring.")
                    return
                
                session.full_name = f"{session.full_name} {last_name}"
                session.step = "waiting_contact"
                
                logger.info(f"✅ User {telegram_id} entered last name: {last_name}")
                
                await message.answer(
                    "📱 <b>Endi kontaktingizni ulashing:</b>\n\n"
                    "Pastdagi tugmani bosib, o'zingizning telefon raqamingizni baham ko'ling.",
                    reply_markup=ReplyKeyboardMarkup(
                        keyboard=[
                            [KeyboardButton(text="📱 Kontaktni ulashish", request_contact=True)],
                        ],
                        resize_keyboard=True,
                        one_time_keyboard=True,
                    ),
                    parse_mode="HTML"
                )
                
            except Exception as exc:
                logger.exception(f"❌ Familiya kiritishda xato: {exc}")
                await message.answer("❌ Xatolik yuz berdi. Iltimos, qayta urinib ko'ring.")
            return  # IMPORTANT: Return to prevent further processing
        
        # ====== OTHER TEXT MESSAGES (MENU HANDLING) ======
        # Continue with existing menu handlers
        # (This handler will now delegate to other handlers if step != waiting_first_name/last_name)

    @dp.message(F.contact)
    async def contact_handler(message: Message) -> None:
        """Handle contact sharing during registration"""
        telegram_id = message.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        logger.info(f"📱 Contact received from user {telegram_id}, step={session.step}")
        
        if session.step != "waiting_contact":
            logger.warning(f"⚠️ User {telegram_id} sent contact at wrong step: {session.step}")
            await message.answer(
                "❌ Kontakt tugmasi bu vaqtda mumkin emas. Iltimos, /start buyrug'ini yuboring.",
                reply_markup=ReplyKeyboardRemove(),
            )
            return
        
        try:
            # Extract phone number from contact
            phone_number = message.contact.phone_number
            
            if not phone_number:
                logger.warning(f"⚠️ User {telegram_id} sent contact without phone number")
                await message.answer(
                    "❌ Telefon raqami topilmadi. Iltimos, kontaktni qayta ulashga urinib ko'ring.",
                    reply_markup=ReplyKeyboardMarkup(
                        keyboard=[
                            [KeyboardButton(text="📱 Kontaktni ulashish", request_contact=True)],
                        ],
                        resize_keyboard=True,
                        one_time_keyboard=True,
                    ),
                )
                return
            
            # ========== SECURITY: Validate Phone Number ==========
            if not input_validator.validate_phone(phone_number):
                logger.warning(f"🚨 Invalid phone format from user {telegram_id}: {phone_number}")
                await message.answer("❌ Telefon raqami noto'g'ri formatda. Iltimos, qayta urinib ko'ring.")
                return
            
            session.phone_number = phone_number
            user_email = f"{telegram_id}@telegram.local"
            session.email = user_email
            
            logger.info(f"✅ Creating user: {telegram_id}, name={session.full_name}")
            
            # Create user in Supabase
            result = await ctx.supabase_service.create_user(
                telegram_id=telegram_id,
                full_name=session.full_name,
                email=user_email,
                phone_number=phone_number,
            )
            
            if result:
                session.step = "ready"
                logger.info(f"✅ User {telegram_id} registered successfully!")
                
                # Send confirmation
                await message.answer(
                    f"✅ <b>Ro'yxatdan o'tish yakunlandi!</b>\n\n"
                    f"👤 Ism: {session.full_name}\n"
                    f"📱 Telefon: {phone_number}\n"
                    f"🎉 Xush kelibsiz!",
                    reply_markup=ReplyKeyboardRemove(),
                    parse_mode="HTML"
                )
                
                # Show main menu
                await message.answer(
                    "📋 Menyudan kerakli bo'limni tanlang:",
                    reply_markup=build_main_menu(),
                )
            else:
                raise Exception("User creation failed - result is False")
            
        except Exception as exc:
            logger.exception(f"❌ Contact handler error for {telegram_id}: {exc}")
            session.step = "idle"
            await message.answer(
                f"❌ Ro'yxatdan o'tishda xatolik yuz berdi:\n\n{str(exc)}\n\n"
                "Iltimos, /start buyrug'ini yuboring.",
                reply_markup=ReplyKeyboardRemove(),
            )

    @dp.message(F.text == "Bekor qilish")
    async def cancel_registration_handler(message: Message) -> None:
        """Handle cancel during registration"""
        telegram_id = message.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        if session.step in {"waiting_first_name", "waiting_last_name", "waiting_contact"}:
            session.step = "idle"
            session.full_name = None
            session.phone_number = None
            await message.answer(
                "Ro'yxatdan o'tish bekor qilindi.\n\n"
                "Agar qayta ro'yxatdan o'tmoqchi bo'lsangiz, /start buyrug'ini yuboring.",
                reply_markup=ReplyKeyboardRemove(),
            )
        else:
            await message.answer("Buyruq rad etildi.", reply_markup=ReplyKeyboardRemove())

    @dp.callback_query(F.data.startswith("sheet:"))
    async def select_sheet_handler(callback: CallbackQuery) -> None:
        telegram_id = callback.from_user.id
        session = ctx.sessions.get(telegram_id)
        key = callback.data.split("sheet:", 1)[1]
        if key not in session.pending_sheets:
            await callback.answer("Jadval havolasi muddati tugagan. Qayta ulab ko'ring.", show_alert=True)
            return
        selected_sheet_name = session.pending_sheets[key]
        selected_sheet_id = key.split(":", 1)[1]
        if not session.google_credentials_json:
            await callback.answer("Google ruxsati topilmadi. Qayta ulab ko'ring.", show_alert=True)
            return
        try:
            await callback.message.edit_text(
                "⏳ Google Sheets o'qilmoqda... Iltimos kuting..."
            )
            
            credentials = await asyncio.to_thread(
                credentials_from_json, session.google_credentials_json, telegram_id
            )
            all_sheets_data = await asyncio.to_thread(fetch_sheet_rows, credentials, selected_sheet_id)
            
            session.sheet_id = selected_sheet_id
            session.sheet_name = selected_sheet_name
            session.all_sheets_data = all_sheets_data  # Store ALL sheets
            session.sheet_data = []  # Legacy support
            session.excel_data = []
            session.step = "in_chat"
            
            await ctx.supabase_service.save_integration(
                telegram_id, selected_sheet_id, selected_sheet_name
            )

            # ---- Persist workspace to SQLite so it survives restarts ----
            company_id = _workspace_store.save_workspace(
                telegram_id,
                mode="sheets",
                sheet_id=selected_sheet_id,
                sheet_name=selected_sheet_name,
            )
            if company_id > 0:
                for sheet_title, rows in all_sheets_data.items():
                    cache_key = f"sheet:{selected_sheet_id}:{sheet_title}"
                    _workspace_store.save_cache(company_id, cache_key, sheet_title, rows, telegram_id=telegram_id)
                logger.info(f"💾 Sheet workspace persisted to SQLite+Supabase for user {telegram_id}")
            # ---------------------------------------------------------------
            
            # Show summary of all sheets read
            sheet_summary = "📊 Barcha jadvallar o'qildi:\n\n"
            for sheet_name, rows in all_sheets_data.items():
                row_count = len(rows)
                col_count = len(rows[0]) if rows else 0
                sheet_summary += f"📋 {html_escape(sheet_name)}: {row_count} qator, {col_count} ustun\n"
            
            sheet_summary += f"\n✅ Ulandi: {html_escape(selected_sheet_name)}\n"
            sheet_summary += "💬 Endi savolingizni yozing, jadval ma'lumotlari asosida javob beraman."
            
            await callback.message.edit_text(sheet_summary, reply_markup=build_chat_response_keyboard())
            await callback.answer("Google Sheet muvaffaqiyatli ulandi.")
        except Exception as exc:
            logger.exception("Sheet tanlashda xato: %s", exc)
            await callback.answer(
                "Jadvalni ulashda xatolik yuz berdi. Iltimos, qayta urinib ko'ring.",
                show_alert=True,
            )

    @dp.callback_query(F.data.startswith("select_sheet:"))
    async def select_folder_sheet_handler(callback: CallbackQuery) -> None:
        """Handle selecting individual sheets from a folder"""
        telegram_id = callback.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        try:
            idx_str = callback.data.split("select_sheet:", 1)[1]
            idx = int(idx_str)
            
            if idx >= len(session.folder_spreadsheets):
                await callback.answer("Jadval topilmadi. Qayta ulab ko'ring.", show_alert=True)
                return
            
            sheet_id = session.folder_spreadsheets[idx]['id']
            
            # Toggle selection
            if sheet_id in session.selected_spreadsheets:
                session.selected_spreadsheets.remove(sheet_id)
                status = "❌ Olib tashlandi"
            else:
                session.selected_spreadsheets.append(sheet_id)
                status = "✅ Tanlandi"
            
            await callback.answer(f"{status}: {session.folder_spreadsheets[idx]['name']}")
            
        except Exception as exc:
            logger.exception(f"❌ Folder sheet selection error: {exc}")
            await callback.answer("❌ Xatolik yuz berdi!", show_alert=True)

    @dp.callback_query(F.data == "load_folder_sheets")
    async def load_folder_sheets_handler(callback: CallbackQuery) -> None:
        """Load all selected sheets from the folder"""
        telegram_id = callback.from_user.id
        session = ctx.sessions.get(telegram_id)
        
        try:
            if not session.selected_spreadsheets:
                await callback.answer("Hech qanday spreadsheet tanlanmadi. Iltimos, bitta-bitta tanlang!", show_alert=True)
                return
            
            await callback.message.edit_text(
                f"⏳ {len(session.selected_spreadsheets)} ta spreadsheet o'qilmoqda...\n\n"
                "Bu biroz vaqt olishi mumkin, iltimos kuting..."
            )
            
            # Get credentials
            try:
                creds = credentials_from_json(session.google_credentials_json, telegram_id=telegram_id)
            except ValueError as scope_error:
                logger.info(f"Old scopes detected for user {telegram_id}: {scope_error}")
                await callback.message.edit_text(
                    "🔄 <b>Yangi ruxsatlar kerak</b>\n\n"
                    "Google Sheets/Drive ning naqsh o'zgarib ketgan. Iltimos, qayta avtentifikatsiya qiling.\n\n"
                    "Asosiy menyudan \"📁 Google Drive Folder\" tugmasini bosing.",
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🏠 Asosiy menyuya qaytish", callback_data="main_menu")]])
                )
                session.google_credentials_json = None
                return
            
            # Import the Google Drive service
            from google_drive_service import GoogleDriveService
            
            drive_service = GoogleDriveService(creds)
            
            # Read all selected spreadsheets
            all_folder_sheets_data = {}
            failed_sheets = []
            
            for idx, sheet_id in enumerate(session.selected_spreadsheets, 1):
                try:
                    logger.info(f"📊 Reading spreadsheet {idx}/{len(session.selected_spreadsheets)}: {sheet_id}")
                    
                    sheets_data = await drive_service.read_spreadsheet(sheet_id)
                    all_folder_sheets_data[sheet_id] = sheets_data
                    
                    # Find the name from folder_spreadsheets
                    sheet_name = next(
                        (s['name'] for s in session.folder_spreadsheets if s['id'] == sheet_id),
                        sheet_id
                    )
                    logger.info(f"✅ Successfully read: {sheet_name}")
                    
                except Exception as e:
                    logger.error(f"❌ Failed to read spreadsheet {sheet_id}: {e}")
                    sheet_name = next(
                        (s['name'] for s in session.folder_spreadsheets if s['id'] == sheet_id),
                        sheet_id
                    )
                    failed_sheets.append(sheet_name)
            
            if not all_folder_sheets_data:
                await callback.message.edit_text(
                    "❌ Hech qanday spreadsheet o'qilmadi.\n\n"
                    "💡 Papka havolasini to'g'ri qilib, qayta urinib ko'ring.",
                    reply_markup=build_retry_keyboard("folder")
                )
                session.step = "ready"
                return
            
            # Store the data
            session.all_folder_sheets_data = all_folder_sheets_data
            session.step = "in_chat"

            # ---- Persist workspace to SQLite so it survives restarts ----
            folder_url = None
            if session.folder_id:
                folder_url = f"https://drive.google.com/drive/folders/{session.folder_id}"
            company_id = _workspace_store.save_workspace(
                telegram_id,
                mode="folder",
                folder_id=session.folder_id,
                folder_url=folder_url,
                folder_spreadsheets=session.folder_spreadsheets,
                selected_spreadsheets=session.selected_spreadsheets,
            )
            if company_id > 0:
                for sid, sheets_dict in all_folder_sheets_data.items():
                    for sheet_title, rows in sheets_dict.items():
                        cache_key = f"folder:{sid}:{sheet_title}"
                        _workspace_store.save_cache(company_id, cache_key, sheet_title, rows, telegram_id=telegram_id)
                logger.info(f"💾 Folder workspace persisted to SQLite+Supabase for user {telegram_id}")
            # ---------------------------------------------------------------
            
            # Create summary
            summary = "✅ <b>Google Drive spreadsheetlari muvaffaqiyatli ulandi!</b>\n\n"
            summary += f"📊 <b>{len(all_folder_sheets_data)} ta spreadsheet o'qildi:</b>\n\n"
            
            for sheet_id, sheets in all_folder_sheets_data.items():
                sheet_name = next(
                    (s['name'] for s in session.folder_spreadsheets if s['id'] == sheet_id),
                    "Unknown"
                )
                summary += f"📁 <b>{html_escape(sheet_name)}</b>\n"
                for sheet_title, rows in sheets.items():
                    row_count = len(rows)
                    col_count = len(rows[0]) if rows else 0
                    summary += f"   └─ 📋 {html_escape(sheet_title)}: {row_count} qator, {col_count} ustun\n"
                summary += "\n"
            
            if failed_sheets:
                summary += f"\n⚠️ <b>O'qilmagan spreadsheetlar:</b>\n"
                for name in failed_sheets:
                    summary += f"   ❌ {html_escape(name)}\n"
            
            summary += "\n💬 Endi savolingizni yozing, barcha spreadsheet ma'lumotlari asosida javob beraman."
            
            try:
                await callback.message.edit_text(
                    summary,
                    parse_mode="HTML",
                    reply_markup=build_chat_response_keyboard()
                )
            except Exception:
                # Fallback: send without HTML if parsing fails
                plain_summary = summary.replace("<b>", "").replace("</b>", "")
                await callback.message.edit_text(
                    plain_summary,
                    reply_markup=build_chat_response_keyboard()
                )
            
            logger.info(f"✅ Successfully loaded {len(all_folder_sheets_data)} spreadsheets for user {telegram_id}")
            
        except Exception as exc:
            logger.exception(f"❌ Error loading folder sheets: {exc}")
            await callback.message.edit_text(
                "❌ Spreadsheetlari yuklashda xatolik yuz berdi.\n\n"
                "💡 Qayta urinib ko'ring.",
                reply_markup=build_retry_keyboard("folder")
            )
            session.step = "ready"

    @dp.message(F.document)
    async def document_handler(message: Message, bot: Bot) -> None:
        telegram_id = message.from_user.id
        
        # ========== SECURITY: Rate Limiting ==========
        if not rate_limiter.is_allowed(telegram_id):
            await message.answer(
                "⏸️ <b>Juda ko'p so'rovlar!</b>\n\n"
                "Iltimos, bir daqiqa kutib turing va qayta urinib ko'ring.",
                parse_mode="HTML"
            )
            return
        
        session = ctx.sessions.get(telegram_id)
        if session.step != "waiting_excel":
            await message.answer(
                "Agar Excel yuklamoqchi bo'lsangiz, avval menyudan \"📁 Excel fayl yuklash\" tugmasini bosing."
            )
            return

        doc = message.document
        file_name = doc.file_name or ""
        
        # ========== SECURITY: Validate File ==========
        if not file_name:
            await message.answer("❌ Fayl nomi topilmadi.")
            return
        
        # Sanitize filename
        file_name = FileValidator.sanitize_filename(file_name)
        
        # Check file extension
        if not file_name.lower().endswith((".xlsx", ".xls", ".xlsm")):
            await message.answer(
                "❌ Noto'g'ri fayl turi.\n\n"
                "Faqat quyidagi formatlar ruxsat:\n"
                "• .xlsx\n"
                "• .xls\n"
                "• .xlsm"
            )
            return

        try:
            # Download file
            telegram_file = await bot.get_file(doc.file_id)
            buffer = io.BytesIO()
            await bot.download_file(telegram_file.file_path, destination=buffer)
            file_content = buffer.getvalue()
            
            # ========== SECURITY: Validate Excel File ==========
            is_valid, error_msg = FileValidator.validate_excel_file(file_name, file_content)
            if not is_valid:
                logger.warning(f"🚨 File validation failed for user {telegram_id}: {error_msg}")
                await message.answer(error_msg)
                session.step = "ready"
                return
            
            # Parse Excel
            excel_rows = await asyncio.to_thread(parse_excel_bytes, file_name, file_content)
            if not excel_rows:
                await message.answer("❌ Fayl bo'sh ko'rinmoqda. Iltimos, boshqa fayl yuboring.")
                return
            
            session.excel_data = excel_rows
            session.sheet_data = []
            session.sheet_id = None
            session.sheet_name = file_name
            session.step = "in_chat"
            
            logger.info(f"✅ Excel file loaded for user {telegram_id}: {len(excel_rows)} rows")
            
            await message.answer(
                f"✅ <b>Excel fayl muvaffaqiyatli yuklandi!</b>\n\n"
                f"📊 Qatorlar: {len(excel_rows)}\n"
                f"📄 Fayl: {file_name}\n\n"
                f"💬 Endi savolingizni yozing, jadval ma'lumotlari asosida javob beraman.",
                parse_mode="HTML",
                reply_markup=build_chat_response_keyboard()
            )
        except Exception as exc:
            logger.exception(f"❌ Excel file processing error for user {telegram_id}: {exc}")
            await message.answer(
                "❌ Excel faylni o'qishda xatolik yuz berdi.\n\n"
                "Sabablari:\n"
                "• Fayl korruptsiyalangan\n"
                "• Fayl turi to'g'ri emas\n"
                "• Fayl juda katta\n\n"
                "Iltimos, boshqa fayl yuboring yoki /start buyrug'ini yuboring."
            )
            session.step = "ready"


    @dp.message(F.text)
    async def text_router(message: Message) -> None:
        telegram_id = message.from_user.id
        session = ctx.sessions.get(telegram_id)
        text = (message.text or "").strip()
        if not text:
            return

        if text in {MAIN_MENU_SHEETS, MAIN_MENU_EXCEL}:
            return

        if session.step == "waiting_magic_link":
            # User should click the magic link button instead
            await message.answer("Ro'yxatdan o'tish uchun `/start` buyrug'ini yuboring.")
            return

        if session.step != "ready":
            await message.answer("Jarayonni boshlash uchun /start buyrug'ini yuboring.")
            return

        context_rows = session.sheet_data or session.excel_data
        if not context_rows:
            await message.answer(
                "Avval Google Sheets ulang yoki Excel fayl yuklang, keyin savol bering."
            )
            return

        await message.answer("Savolingiz qabul qilindi, javob tayyorlanmoqda...")
        context_text = table_to_text(context_rows)
        
        try:
            # Generate response based on available context data
            answer = (
                f"📊 *Sizning savolingizga javob:*\n\n"
                f"Savolingiz: {text}\n\n"
                f"📋 *Jadval ma'lumoti:*\n{context_text[:1500]}\n\n"
                f"Bu javob berilgan jadval ma'lumotlariga tayanadi."
            )
            
            if not answer:
                answer = "Kechirasiz, hozircha aniq javob hosil bo'lmadi. Iltimos, savolni boshqacha yuboring."
            await message.answer(answer, parse_mode="Markdown")
            await ctx.supabase_service.save_message(
                telegram_id=telegram_id,
                question=text,
                answer=answer,
            )
        except Exception as exc:
            # Handle any errors
            error_msg = str(exc).lower()
            if "rate" in error_msg or "too many" in error_msg:
                await message.answer(
                    "So'rovlar soni juda ko'p. Iltimos, birozdan so'ng qayta urinib ko'ring."
                )
            elif "connection" in error_msg or "network" in error_msg:
                await message.answer(
                    "Tarmoq bilan bog'liq xatolik yuz berdi. Internetni tekshirib, qayta urinib ko'ring."
                )
            else:
                await message.answer(
                    "Xizmatda vaqtinchalik xatolik yuz berdi. Iltimos, keyinroq urinib ko'ring."
                )
                logger.exception("Javobida xato: %s", exc)


async def main() -> None:
    config = Config.from_env()
    context = AppContext(config)
    
    # Create bot with default settings
    bot = Bot(
        token=config.bot_token,
        default=DefaultBotProperties(parse_mode=ParseMode.HTML),
    )
    context.bot = bot
    
    # Ensure any previous bot sessions are closed
    logger.info("🔄 Cleaning up any previous bot sessions...")
    await asyncio.sleep(2)  # Give Telegram time to register session termination
    
    # Test the bot token first
    try:
        me = await bot.get_me()
        logger.info(f"Bot authenticated: {me.first_name} (@{me.username})")
    except Exception as e:
        logger.error(f"Failed to authenticate bot: {e}")
        raise
    
    # Register bot commands (shown at the bottom left)
    try:
        from aiogram.types import BotCommand
        commands = [
            BotCommand(command="start", description="🏠 Asosiy menyu - Boshlanish"),
            BotCommand(command="chat", description="💬 Chat - Assistant bilan suhbat"),
            BotCommand(command="help", description="❓ Yordam - Qo'llanma"),
        ]
        await bot.set_my_commands(commands)
        logger.info("✅ Bot commands registered")
    except Exception as e:
        logger.warning(f"⚠️  Could not register bot commands: {e}")
    
    # Force-close any existing polling session by deleting webhook + dropping updates
    # This kills any old bot instance that's still polling
    try:
        logger.info("🔄 Forcefully stopping any existing bot sessions...")
        await bot.delete_webhook(drop_pending_updates=True)
        logger.info("   ✅ Webhook deleted, pending updates dropped")
        # Wait longer for Telegram to fully release the old polling session
        # (needed when migrating from another server like Render)
        await asyncio.sleep(10)
    except Exception as e:
        logger.warning(f"⚠️  Could not clean webhook: {e}")
        await asyncio.sleep(5)
    
    dp = Dispatcher()
    register_handlers(dp, context)

    # OAuth server - REQUIRED for Google Sheets integration
    oauth_server = OAuthServer(context)
    await oauth_server.start()
    
    logger.info("Starting bot polling loop...")
    polling_task = None
    try:
        # Start polling - aiogram handles reconnection automatically
        polling_task = asyncio.create_task(
            dp.start_polling(
                bot,
                allowed_updates=dp.resolve_used_update_types(),
                relax_timeout=60.0,
                long_poll_timeout=60.0,
            )
        )
        await polling_task
    except (asyncio.CancelledError, KeyboardInterrupt):
        logger.info("Polling cancelled")
        if polling_task and not polling_task.done():
            polling_task.cancel()
            try:
                await polling_task
            except asyncio.CancelledError:
                pass
    except Exception as e:
        logger.error(f"Polling error: {type(e).__name__}: {e}")
    finally:
        await oauth_server.stop()  # Stop OAuth server
        try:
            await bot.session.close()
        except:
            pass


if __name__ == "__main__":
    logger.info("🚀 OnBrain AI Bot Starting...")
    while True:
        try:
            asyncio.run(main())
            logger.info("Bot run completed normally")
        except KeyboardInterrupt:
            logger.info("Bot stopped by user (Ctrl+C)")
            break
        except asyncio.CancelledError:
            logger.warning("Bot was cancelled, restarting...")
            time.sleep(10)
        except Exception as e:
            error_name = type(e).__name__
            logger.error(f"Bot error ({error_name}): {e}")
            if "conflict" in str(e).lower() or "terminated by other" in str(e).lower():
                logger.info("⚠️ Another bot instance detected. Waiting 15s for it to stop...")
                time.sleep(15)
            else:
                logger.info("Restarting bot in 5 seconds...")
                time.sleep(5)
